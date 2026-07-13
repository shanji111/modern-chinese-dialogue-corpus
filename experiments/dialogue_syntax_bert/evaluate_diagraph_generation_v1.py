from __future__ import annotations

import argparse
import csv
import json
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
ARTIFACT_ROOT = ROOT / "artifacts" / "formal_300_v1"
GOLD_DIR = ARTIFACT_ROOT / "diagraph_gold_50" / "gold_v1"
PAIR_LIST_PATH = ARTIFACT_ROOT / "diagraph_gold_50" / "diagraph_gold_50_pair_list.csv"
EVALUATOR_DIR = ARTIFACT_ROOT / "diagraph_generation_evaluation_v1" / "evaluator_v1"

DEFAULT_GOLD_ACTIVE = GOLD_DIR / "diagraph_gold_50_column_gold_v1_active.csv"
DEFAULT_TOY_PREDICTION = EVALUATOR_DIR / "toy_prediction_exact_subset.csv"
DEFAULT_TOY_RUN_DIR = EVALUATOR_DIR / "toy_run"
DEFAULT_IMPLEMENTATION_REPORT = (
    EVALUATOR_DIR / "diagraph_evaluation_runner_v1_implementation_report.md"
)

GOLD_REQUIRED_FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
]
PRED_REQUIRED_FIELDS = [
    "annotation_id",
    "pair_id",
    "pred_column_id",
    "pred_span_a",
    "pred_span_b",
    "pred_relation_type",
    "pred_relation_strength",
    "pred_alignment_direction",
    "pred_is_core_column",
    "pred_supports_resonance",
    "pred_confidence",
    "generator_name",
    "generator_version",
    "notes",
]
PAIR_REQUIRED_FIELDS = [
    "annotation_id",
    "pair_id",
    "source",
    "dataset_name",
    "sample_stratum",
    "turn_a",
    "turn_b",
]

VALID_RELATION_TYPES = {
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
}
VALID_RELATION_STRENGTHS = {"strong", "medium", "weak"}
VALID_ALIGNMENT_DIRECTIONS = {"A_to_B", "B_to_A", "mutual"}
VALID_BINARY_OR_UNKNOWN = {"0", "1", "?"}


@dataclass(frozen=True)
class ColumnMatch:
    annotation_id: str
    pair_id: str
    gold_column_id: str
    pred_column_id: str
    gold_span_a: str
    pred_span_a: str
    gold_span_b: str
    pred_span_b: str
    gold_relation_type: str
    pred_relation_type: str
    gold_relation_strength: str
    pred_relation_strength: str
    gold_is_core_column: str
    pred_is_core_column: str
    gold_supports_resonance: str
    pred_supports_resonance: str
    gold_alignment_direction: str
    pred_alignment_direction: str
    span_a_overlap_ratio: float
    span_b_overlap_ratio: float
    mean_overlap_ratio: float
    match_kind: str


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: Iterable[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(
    path: Path,
    rows: List[Dict[str, object]],
    fieldnames: Sequence[str],
    sheet_name: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "Sheet1"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, fieldname in enumerate(fieldnames, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=fieldname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, fieldname in enumerate(fieldnames, start=1):
            value = row.get(fieldname, "")
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, fieldname in enumerate(fieldnames, start=1):
        values = [str(fieldname)]
        for row in rows:
            value = row.get(fieldname, "")
            values.append("" if value is None else str(value))
        width = min(max(len(value) for value in values) + 2, 48)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 10)

    workbook.save(path)


def ensure_fields(rows: List[Dict[str, str]], required_fields: Sequence[str], name: str) -> None:
    if not rows:
        raise ValueError(f"{name} is empty.")
    missing = [field for field in required_fields if field not in rows[0]]
    if missing:
        raise ValueError(f"{name} missing required fields: {missing}")


def safe_divide(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


def f1(precision: float, recall: float) -> float:
    return 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0


def normalize_boolish(value: str) -> str:
    return (value or "").strip()


def pair_key(row: Dict[str, str]) -> str:
    return row["annotation_id"]


def gold_key(row: Dict[str, str]) -> str:
    return f"{row['annotation_id']}/{row['column_id']}"


def pred_key(row: Dict[str, str]) -> str:
    return f"{row['annotation_id']}/{row['pred_column_id']}"


def longest_common_substring_len(a: str, b: str) -> int:
    if not a or not b:
        return 0
    prev = [0] * (len(b) + 1)
    best = 0
    for ca in a:
        current = [0] * (len(b) + 1)
        for idx, cb in enumerate(b, start=1):
            if ca == cb:
                current[idx] = prev[idx - 1] + 1
                if current[idx] > best:
                    best = current[idx]
        prev = current
    return best


def span_overlap_ratio(a: str, b: str) -> float:
    a = a or ""
    b = b or ""
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    shorter = min(len(a), len(b))
    if shorter == 0:
        return 0.0
    if a in b or b in a:
        return 1.0
    overlap = longest_common_substring_len(a, b)
    return overlap / shorter


def format_metric(value: float) -> float:
    return round(value, 6)


def build_pair_map(pair_rows: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    ensure_fields(pair_rows, PAIR_REQUIRED_FIELDS, "pair_list")
    return {row["annotation_id"]: row for row in pair_rows}


def validate_gold_rows(gold_rows: List[Dict[str, str]], pair_map: Dict[str, Dict[str, str]]) -> None:
    ensure_fields(gold_rows, GOLD_REQUIRED_FIELDS, "gold_active")
    duplicates: set[str] = set()
    seen: set[str] = set()
    missing_pairs: List[str] = []
    invalid_rows: List[str] = []
    for row in gold_rows:
        key = gold_key(row)
        if key in seen:
            duplicates.add(key)
        seen.add(key)
        if row["annotation_id"] not in pair_map:
            missing_pairs.append(key)
        if row["relation_type"] not in VALID_RELATION_TYPES:
            invalid_rows.append(key)
        if row["relation_strength"] not in VALID_RELATION_STRENGTHS:
            invalid_rows.append(key)
        if row["alignment_direction"] not in VALID_ALIGNMENT_DIRECTIONS:
            invalid_rows.append(key)
        if row["is_core_column"] not in {"0", "1"} or row["supports_resonance"] not in {"0", "1"}:
            invalid_rows.append(key)
    if duplicates:
        raise ValueError(f"gold_active has duplicate annotation_id/column_id keys: {sorted(duplicates)}")
    if missing_pairs:
        raise ValueError(f"gold_active contains annotation_ids missing from pair_list: {missing_pairs[:10]}")
    if invalid_rows:
        raise ValueError(f"gold_active contains invalid value-domain rows: {invalid_rows[:10]}")


def validate_prediction_rows(
    prediction_rows: List[Dict[str, str]],
    gold_annotation_ids: set[str],
    pair_map: Dict[str, Dict[str, str]],
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    ensure_fields(prediction_rows, PRED_REQUIRED_FIELDS, "prediction")
    valid_rows: List[Dict[str, str]] = []
    invalid_rows: List[Dict[str, str]] = []

    for row in prediction_rows:
        reasons: List[str] = []
        annotation_id = row["annotation_id"]
        pair_id = row["pair_id"]
        pair = pair_map.get(annotation_id)
        if annotation_id not in gold_annotation_ids:
            reasons.append("annotation_id_not_in_gold_50")
        if pair is None:
            reasons.append("annotation_id_missing_from_pair_list")
        else:
            if pair_id != pair["pair_id"]:
                reasons.append("pair_id_mismatch")
            if row["pred_span_a"] not in pair["turn_a"]:
                reasons.append("pred_span_a_not_found_in_turn_a")
            if row["pred_span_b"] not in pair["turn_b"]:
                reasons.append("pred_span_b_not_found_in_turn_b")
        if row["pred_relation_type"] not in VALID_RELATION_TYPES:
            reasons.append("invalid_pred_relation_type")
        if row["pred_relation_strength"] not in VALID_RELATION_STRENGTHS:
            reasons.append("invalid_pred_relation_strength")
        if row["pred_alignment_direction"] not in VALID_ALIGNMENT_DIRECTIONS:
            reasons.append("invalid_pred_alignment_direction")
        if normalize_boolish(row["pred_is_core_column"]) not in VALID_BINARY_OR_UNKNOWN:
            reasons.append("invalid_pred_is_core_column")
        if normalize_boolish(row["pred_supports_resonance"]) not in VALID_BINARY_OR_UNKNOWN:
            reasons.append("invalid_pred_supports_resonance")
        if not row["pred_span_a"]:
            reasons.append("empty_pred_span_a")
        if not row["pred_span_b"]:
            reasons.append("empty_pred_span_b")

        if reasons:
            invalid_row = dict(row)
            invalid_row["invalid_reason"] = "|".join(reasons)
            invalid_rows.append(invalid_row)
        else:
            valid_rows.append(dict(row))

    return valid_rows, invalid_rows


def compute_exact_matches(
    gold_rows: List[Dict[str, str]],
    pred_rows: List[Dict[str, str]],
) -> Tuple[List[ColumnMatch], set[str], set[str]]:
    gold_by_pair: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    pred_by_pair: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in gold_rows:
        gold_by_pair[row["annotation_id"]].append(row)
    for row in pred_rows:
        pred_by_pair[row["annotation_id"]].append(row)

    matches: List[ColumnMatch] = []
    matched_gold: set[str] = set()
    matched_pred: set[str] = set()

    for annotation_id in sorted(gold_by_pair):
        pair_gold = sorted(gold_by_pair[annotation_id], key=lambda item: item["column_id"])
        pair_pred = sorted(
            pred_by_pair.get(annotation_id, []),
            key=lambda item: item["pred_column_id"],
        )
        for gold in pair_gold:
            for pred in pair_pred:
                gk = gold_key(gold)
                pk = pred_key(pred)
                if gk in matched_gold or pk in matched_pred:
                    continue
                if (
                    gold["span_a"] == pred["pred_span_a"]
                    and gold["span_b"] == pred["pred_span_b"]
                ):
                    matched_gold.add(gk)
                    matched_pred.add(pk)
                    matches.append(
                        ColumnMatch(
                            annotation_id=annotation_id,
                            pair_id=gold["pair_id"],
                            gold_column_id=gold["column_id"],
                            pred_column_id=pred["pred_column_id"],
                            gold_span_a=gold["span_a"],
                            pred_span_a=pred["pred_span_a"],
                            gold_span_b=gold["span_b"],
                            pred_span_b=pred["pred_span_b"],
                            gold_relation_type=gold["relation_type"],
                            pred_relation_type=pred["pred_relation_type"],
                            gold_relation_strength=gold["relation_strength"],
                            pred_relation_strength=pred["pred_relation_strength"],
                            gold_is_core_column=gold["is_core_column"],
                            pred_is_core_column=pred["pred_is_core_column"],
                            gold_supports_resonance=gold["supports_resonance"],
                            pred_supports_resonance=pred["pred_supports_resonance"],
                            gold_alignment_direction=gold["alignment_direction"],
                            pred_alignment_direction=pred["pred_alignment_direction"],
                            span_a_overlap_ratio=1.0,
                            span_b_overlap_ratio=1.0,
                            mean_overlap_ratio=1.0,
                            match_kind="exact",
                        )
                    )
                    break

    return matches, matched_gold, matched_pred


def compute_relaxed_matches(
    gold_rows: List[Dict[str, str]],
    pred_rows: List[Dict[str, str]],
    relaxed_threshold: float,
) -> List[ColumnMatch]:
    exact_matches, matched_gold, matched_pred = compute_exact_matches(gold_rows, pred_rows)
    gold_lookup = {gold_key(row): row for row in gold_rows}
    pred_lookup = {pred_key(row): row for row in pred_rows}

    candidate_rows: List[Tuple[float, float, float, str, str, str]] = []
    for gold in gold_rows:
        gk = gold_key(gold)
        if gk in matched_gold:
            continue
        for pred in pred_rows:
            pk = pred_key(pred)
            if pk in matched_pred:
                continue
            if gold["annotation_id"] != pred["annotation_id"]:
                continue
            overlap_a = span_overlap_ratio(gold["span_a"], pred["pred_span_a"])
            overlap_b = span_overlap_ratio(gold["span_b"], pred["pred_span_b"])
            if overlap_a >= relaxed_threshold and overlap_b >= relaxed_threshold:
                mean_overlap = (overlap_a + overlap_b) / 2
                candidate_rows.append(
                    (
                        mean_overlap,
                        overlap_a,
                        overlap_b,
                        gk,
                        pk,
                        gold["annotation_id"],
                    )
                )

    candidate_rows.sort(
        key=lambda item: (
            -item[0],
            -item[1],
            -item[2],
            item[5],
            item[3],
            item[4],
        )
    )

    relaxed_only: List[ColumnMatch] = []
    for mean_overlap, overlap_a, overlap_b, gk, pk, _annotation_id in candidate_rows:
        if gk in matched_gold or pk in matched_pred:
            continue
        gold = gold_lookup[gk]
        pred = pred_lookup[pk]
        matched_gold.add(gk)
        matched_pred.add(pk)
        relaxed_only.append(
            ColumnMatch(
                annotation_id=gold["annotation_id"],
                pair_id=gold["pair_id"],
                gold_column_id=gold["column_id"],
                pred_column_id=pred["pred_column_id"],
                gold_span_a=gold["span_a"],
                pred_span_a=pred["pred_span_a"],
                gold_span_b=gold["span_b"],
                pred_span_b=pred["pred_span_b"],
                gold_relation_type=gold["relation_type"],
                pred_relation_type=pred["pred_relation_type"],
                gold_relation_strength=gold["relation_strength"],
                pred_relation_strength=pred["pred_relation_strength"],
                gold_is_core_column=gold["is_core_column"],
                pred_is_core_column=pred["pred_is_core_column"],
                gold_supports_resonance=gold["supports_resonance"],
                pred_supports_resonance=pred["pred_supports_resonance"],
                gold_alignment_direction=gold["alignment_direction"],
                pred_alignment_direction=pred["pred_alignment_direction"],
                span_a_overlap_ratio=overlap_a,
                span_b_overlap_ratio=overlap_b,
                mean_overlap_ratio=mean_overlap,
                match_kind="relaxed",
            )
        )

    return exact_matches + relaxed_only


def build_match_rows(matches: List[ColumnMatch]) -> List[Dict[str, object]]:
    return [
        {
            "annotation_id": match.annotation_id,
            "pair_id": match.pair_id,
            "gold_column_id": match.gold_column_id,
            "pred_column_id": match.pred_column_id,
            "gold_span_a": match.gold_span_a,
            "pred_span_a": match.pred_span_a,
            "gold_span_b": match.gold_span_b,
            "pred_span_b": match.pred_span_b,
            "gold_relation_type": match.gold_relation_type,
            "pred_relation_type": match.pred_relation_type,
            "gold_relation_strength": match.gold_relation_strength,
            "pred_relation_strength": match.pred_relation_strength,
            "gold_is_core_column": match.gold_is_core_column,
            "pred_is_core_column": match.pred_is_core_column,
            "gold_supports_resonance": match.gold_supports_resonance,
            "pred_supports_resonance": match.pred_supports_resonance,
            "gold_alignment_direction": match.gold_alignment_direction,
            "pred_alignment_direction": match.pred_alignment_direction,
            "span_a_overlap_ratio": format_metric(match.span_a_overlap_ratio),
            "span_b_overlap_ratio": format_metric(match.span_b_overlap_ratio),
            "mean_overlap_ratio": format_metric(match.mean_overlap_ratio),
            "match_kind": match.match_kind,
            "relation_type_match": int(match.gold_relation_type == match.pred_relation_type),
            "alignment_direction_match": int(
                match.gold_alignment_direction == match.pred_alignment_direction
            ),
            "core_match": int(
                match.gold_is_core_column == "1" and match.pred_is_core_column == "1"
            ),
        }
        for match in matches
    ]


def build_per_pair_metrics(
    gold_rows: List[Dict[str, str]],
    valid_pred_rows: List[Dict[str, str]],
    invalid_pred_rows: List[Dict[str, str]],
    exact_matches: List[ColumnMatch],
    relaxed_matches: List[ColumnMatch],
) -> List[Dict[str, object]]:
    gold_by_pair: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    valid_pred_by_pair: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    invalid_pred_by_pair: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    exact_by_pair: Dict[str, List[ColumnMatch]] = defaultdict(list)
    relaxed_by_pair: Dict[str, List[ColumnMatch]] = defaultdict(list)

    for row in gold_rows:
        gold_by_pair[row["annotation_id"]].append(row)
    for row in valid_pred_rows:
        valid_pred_by_pair[row["annotation_id"]].append(row)
    for row in invalid_pred_rows:
        invalid_pred_by_pair[row["annotation_id"]].append(row)
    for match in exact_matches:
        exact_by_pair[match.annotation_id].append(match)
    for match in relaxed_matches:
        relaxed_by_pair[match.annotation_id].append(match)

    metrics_rows: List[Dict[str, object]] = []
    for annotation_id in sorted(gold_by_pair):
        gold_pair = gold_by_pair[annotation_id]
        pair_id = gold_pair[0]["pair_id"]
        gold_count = len(gold_pair)
        valid_pred_count = len(valid_pred_by_pair.get(annotation_id, []))
        invalid_pred_count = len(invalid_pred_by_pair.get(annotation_id, []))
        exact_count = len(exact_by_pair.get(annotation_id, []))
        relaxed_count = len(relaxed_by_pair.get(annotation_id, []))
        overgenerated_count = valid_pred_count - relaxed_count
        missing_gold_count = gold_count - relaxed_count
        gold_core_count = sum(1 for row in gold_pair if row["is_core_column"] == "1")
        pred_core_count = sum(
            1
            for row in valid_pred_by_pair.get(annotation_id, [])
            if normalize_boolish(row["pred_is_core_column"]) == "1"
        )
        matched_core_count = sum(
            1
            for match in relaxed_by_pair.get(annotation_id, [])
            if match.gold_is_core_column == "1" and normalize_boolish(match.pred_is_core_column) == "1"
        )
        missing_core_count = gold_core_count - matched_core_count
        exact_precision = safe_divide(exact_count, valid_pred_count)
        exact_recall = safe_divide(exact_count, gold_count)
        relaxed_precision = safe_divide(relaxed_count, valid_pred_count)
        relaxed_recall = safe_divide(relaxed_count, gold_count)
        high_overgeneration = int(
            overgenerated_count >= 2 and overgenerated_count > max(1, gold_count // 2)
        )
        metrics_rows.append(
            {
                "annotation_id": annotation_id,
                "pair_id": pair_id,
                "gold_column_count": gold_count,
                "valid_pred_column_count": valid_pred_count,
                "invalid_pred_column_count": invalid_pred_count,
                "exact_match_count": exact_count,
                "relaxed_match_count": relaxed_count,
                "exact_column_precision": format_metric(exact_precision),
                "exact_column_recall": format_metric(exact_recall),
                "exact_column_f1": format_metric(f1(exact_precision, exact_recall)),
                "relaxed_column_precision": format_metric(relaxed_precision),
                "relaxed_column_recall": format_metric(relaxed_recall),
                "relaxed_column_f1": format_metric(f1(relaxed_precision, relaxed_recall)),
                "gold_core_count": gold_core_count,
                "pred_core_count": pred_core_count,
                "matched_core_count": matched_core_count,
                "missing_core_count": missing_core_count,
                "overgenerated_count": overgenerated_count,
                "missing_gold_count": missing_gold_count,
                "column_count_abs_error": abs(valid_pred_count - gold_count),
                "zero_matched_columns": int(relaxed_count == 0),
                "missing_core_columns": int(missing_core_count > 0),
                "high_overgeneration": high_overgeneration,
            }
        )
    return metrics_rows


def build_confusion_matrix(matches: List[ColumnMatch]) -> List[Dict[str, object]]:
    all_types = sorted(VALID_RELATION_TYPES)
    gold_to_pred_counts: Dict[str, Counter[str]] = defaultdict(Counter)
    for match in matches:
        gold_to_pred_counts[match.gold_relation_type][match.pred_relation_type] += 1
    rows: List[Dict[str, object]] = []
    for gold_type in all_types:
        row: Dict[str, object] = {"gold_relation_type": gold_type}
        total = 0
        for pred_type in all_types:
            count = gold_to_pred_counts[gold_type][pred_type]
            row[pred_type] = count
            total += count
        row["row_total"] = total
        rows.append(row)
    return rows


def build_core_error_report(
    gold_rows: List[Dict[str, str]],
    valid_pred_rows: List[Dict[str, str]],
    relaxed_matches: List[ColumnMatch],
) -> List[Dict[str, object]]:
    relaxed_by_gold = {f"{m.annotation_id}/{m.gold_column_id}": m for m in relaxed_matches}
    relaxed_by_pred = {f"{m.annotation_id}/{m.pred_column_id}": m for m in relaxed_matches}
    rows: List[Dict[str, object]] = []

    for gold in gold_rows:
        if gold["is_core_column"] != "1":
            continue
        gk = f"{gold['annotation_id']}/{gold['column_id']}"
        match = relaxed_by_gold.get(gk)
        if match is None:
            rows.append(
                {
                    "error_type": "missing_gold_core",
                    "annotation_id": gold["annotation_id"],
                    "pair_id": gold["pair_id"],
                    "gold_column_id": gold["column_id"],
                    "pred_column_id": "",
                    "gold_relation_type": gold["relation_type"],
                    "pred_relation_type": "",
                    "gold_is_core_column": gold["is_core_column"],
                    "pred_is_core_column": "",
                    "details": "gold core column unmatched after relaxed matching",
                }
            )
        elif normalize_boolish(match.pred_is_core_column) != "1":
            rows.append(
                {
                    "error_type": "gold_core_predicted_noncore",
                    "annotation_id": gold["annotation_id"],
                    "pair_id": gold["pair_id"],
                    "gold_column_id": gold["column_id"],
                    "pred_column_id": match.pred_column_id,
                    "gold_relation_type": gold["relation_type"],
                    "pred_relation_type": match.pred_relation_type,
                    "gold_is_core_column": gold["is_core_column"],
                    "pred_is_core_column": match.pred_is_core_column,
                    "details": "gold core matched by pred column but pred_is_core_column != 1",
                }
            )

    for pred in valid_pred_rows:
        if normalize_boolish(pred["pred_is_core_column"]) != "1":
            continue
        pk = f"{pred['annotation_id']}/{pred['pred_column_id']}"
        match = relaxed_by_pred.get(pk)
        if match is None:
            rows.append(
                {
                    "error_type": "false_core_unmatched_prediction",
                    "annotation_id": pred["annotation_id"],
                    "pair_id": pred["pair_id"],
                    "gold_column_id": "",
                    "pred_column_id": pred["pred_column_id"],
                    "gold_relation_type": "",
                    "pred_relation_type": pred["pred_relation_type"],
                    "gold_is_core_column": "",
                    "pred_is_core_column": pred["pred_is_core_column"],
                    "details": "predicted core column unmatched after relaxed matching",
                }
            )
        elif match.gold_is_core_column != "1":
            rows.append(
                {
                    "error_type": "false_core_matched_to_gold_noncore",
                    "annotation_id": pred["annotation_id"],
                    "pair_id": pred["pair_id"],
                    "gold_column_id": match.gold_column_id,
                    "pred_column_id": pred["pred_column_id"],
                    "gold_relation_type": match.gold_relation_type,
                    "pred_relation_type": pred["pred_relation_type"],
                    "gold_is_core_column": match.gold_is_core_column,
                    "pred_is_core_column": pred["pred_is_core_column"],
                    "details": "predicted core column matched a gold non-core column",
                }
            )

    return rows


def build_unmatched_gold_rows(
    gold_rows: List[Dict[str, str]],
    relaxed_matches: List[ColumnMatch],
) -> List[Dict[str, object]]:
    matched_gold_keys = {f"{match.annotation_id}/{match.gold_column_id}" for match in relaxed_matches}
    rows: List[Dict[str, object]] = []
    for gold in gold_rows:
        gk = f"{gold['annotation_id']}/{gold['column_id']}"
        if gk in matched_gold_keys:
            continue
        row = dict(gold)
        row["unmatched_reason"] = "no_relaxed_match"
        rows.append(row)
    return rows


def build_overgenerated_pred_rows(
    valid_pred_rows: List[Dict[str, str]],
    relaxed_matches: List[ColumnMatch],
) -> List[Dict[str, object]]:
    matched_pred_keys = {f"{match.annotation_id}/{match.pred_column_id}" for match in relaxed_matches}
    rows: List[Dict[str, object]] = []
    for pred in valid_pred_rows:
        pk = f"{pred['annotation_id']}/{pred['pred_column_id']}"
        if pk in matched_pred_keys:
            continue
        row = dict(pred)
        row["overgenerated_reason"] = "no_relaxed_match"
        rows.append(row)
    return rows


def compute_summary(
    gold_rows: List[Dict[str, str]],
    valid_pred_rows: List[Dict[str, str]],
    invalid_pred_rows: List[Dict[str, str]],
    exact_matches: List[ColumnMatch],
    relaxed_matches: List[ColumnMatch],
    per_pair_metrics: List[Dict[str, object]],
) -> Dict[str, object]:
    gold_count = len(gold_rows)
    valid_pred_count = len(valid_pred_rows)
    invalid_pred_count = len(invalid_pred_rows)
    exact_count = len(exact_matches)
    relaxed_count = len(relaxed_matches)

    exact_precision = safe_divide(exact_count, valid_pred_count)
    exact_recall = safe_divide(exact_count, gold_count)
    relaxed_precision = safe_divide(relaxed_count, valid_pred_count)
    relaxed_recall = safe_divide(relaxed_count, gold_count)

    relation_exact_accuracy = safe_divide(
        sum(1 for match in exact_matches if match.gold_relation_type == match.pred_relation_type),
        exact_count,
    )
    relation_relaxed_accuracy = safe_divide(
        sum(1 for match in relaxed_matches if match.gold_relation_type == match.pred_relation_type),
        relaxed_count,
    )

    gold_core_total = sum(1 for row in gold_rows if row["is_core_column"] == "1")
    pred_core_total = sum(
        1 for row in valid_pred_rows if normalize_boolish(row["pred_is_core_column"]) == "1"
    )
    correctly_predicted_core = sum(
        1
        for match in relaxed_matches
        if match.gold_is_core_column == "1" and normalize_boolish(match.pred_is_core_column) == "1"
    )
    unmatched_pred_count = valid_pred_count - relaxed_count
    unmatched_gold_count = gold_count - relaxed_count
    abs_errors = [float(row["column_count_abs_error"]) for row in per_pair_metrics]

    summary = {
        "gold_column_count": gold_count,
        "valid_prediction_count": valid_pred_count,
        "invalid_prediction_count": invalid_pred_count,
        "exact_match_count": exact_count,
        "relaxed_match_count": relaxed_count,
        "exact_column_precision": format_metric(exact_precision),
        "exact_column_recall": format_metric(exact_recall),
        "exact_column_f1": format_metric(f1(exact_precision, exact_recall)),
        "relaxed_column_precision": format_metric(relaxed_precision),
        "relaxed_column_recall": format_metric(relaxed_recall),
        "relaxed_column_f1": format_metric(f1(relaxed_precision, relaxed_recall)),
        "relation_type_accuracy_on_exact_matches": format_metric(relation_exact_accuracy),
        "relation_type_accuracy_on_relaxed_matches": format_metric(relation_relaxed_accuracy),
        "core_column_precision": format_metric(
            safe_divide(correctly_predicted_core, pred_core_total)
        ),
        "core_column_recall": format_metric(
            safe_divide(correctly_predicted_core, gold_core_total)
        ),
        "missing_core_rate": format_metric(
            safe_divide(gold_core_total - correctly_predicted_core, gold_core_total)
        ),
        "false_core_rate": format_metric(
            safe_divide(pred_core_total - correctly_predicted_core, pred_core_total)
        ),
        "overgeneration_rate": format_metric(safe_divide(unmatched_pred_count, valid_pred_count)),
        "missing_column_rate": format_metric(safe_divide(unmatched_gold_count, gold_count)),
        "mean_abs_column_count_error_by_pair": format_metric(statistics.mean(abs_errors)),
        "pair_count": len(per_pair_metrics),
        "pairs_with_zero_matched_columns": sum(
            int(row["zero_matched_columns"]) for row in per_pair_metrics
        ),
        "pairs_with_missing_core_columns": sum(
            int(row["missing_core_columns"]) for row in per_pair_metrics
        ),
        "pairs_with_high_overgeneration": sum(
            int(row["high_overgeneration"]) for row in per_pair_metrics
        ),
    }
    return summary


def write_summary_markdown(path: Path, summary: Dict[str, object], run_name: str) -> None:
    lines = [
        f"# diagraph evaluation summary: {run_name}",
        "",
        "## Counts",
        "",
        f"- gold columns: {summary['gold_column_count']}",
        f"- valid predictions: {summary['valid_prediction_count']}",
        f"- invalid predictions: {summary['invalid_prediction_count']}",
        f"- exact matches: {summary['exact_match_count']}",
        f"- relaxed matches: {summary['relaxed_match_count']}",
        "",
        "## Column metrics",
        "",
        f"- exact precision: {summary['exact_column_precision']}",
        f"- exact recall: {summary['exact_column_recall']}",
        f"- exact F1: {summary['exact_column_f1']}",
        f"- relaxed precision: {summary['relaxed_column_precision']}",
        f"- relaxed recall: {summary['relaxed_column_recall']}",
        f"- relaxed F1: {summary['relaxed_column_f1']}",
        "",
        "## Type / core metrics",
        "",
        f"- relation_type accuracy on exact matches: {summary['relation_type_accuracy_on_exact_matches']}",
        f"- relation_type accuracy on relaxed matches: {summary['relation_type_accuracy_on_relaxed_matches']}",
        f"- core column precision: {summary['core_column_precision']}",
        f"- core column recall: {summary['core_column_recall']}",
        f"- missing-core rate: {summary['missing_core_rate']}",
        f"- false-core rate: {summary['false_core_rate']}",
        "",
        "## Generation balance",
        "",
        f"- overgeneration rate: {summary['overgeneration_rate']}",
        f"- missing column rate: {summary['missing_column_rate']}",
        f"- mean abs column count error by pair: {summary['mean_abs_column_count_error_by_pair']}",
        "",
        "## Pair-level diagnostics",
        "",
        f"- pairs with zero matched columns: {summary['pairs_with_zero_matched_columns']}",
        f"- pairs with missing core columns: {summary['pairs_with_missing_core_columns']}",
        f"- pairs with high overgeneration: {summary['pairs_with_high_overgeneration']}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


MATCH_FIELDNAMES = [
    "annotation_id",
    "pair_id",
    "gold_column_id",
    "pred_column_id",
    "gold_span_a",
    "pred_span_a",
    "gold_span_b",
    "pred_span_b",
    "gold_relation_type",
    "pred_relation_type",
    "gold_relation_strength",
    "pred_relation_strength",
    "gold_is_core_column",
    "pred_is_core_column",
    "gold_supports_resonance",
    "pred_supports_resonance",
    "gold_alignment_direction",
    "pred_alignment_direction",
    "span_a_overlap_ratio",
    "span_b_overlap_ratio",
    "mean_overlap_ratio",
    "match_kind",
    "relation_type_match",
    "alignment_direction_match",
    "core_match",
]


def evaluate_predictions(
    gold_path: Path,
    pair_list_path: Path,
    prediction_path: Path,
    output_dir: Path,
    relaxed_threshold: float,
    run_name: str,
) -> Dict[str, object]:
    gold_rows = read_csv_dicts(gold_path)
    pair_rows = read_csv_dicts(pair_list_path)
    prediction_rows = read_csv_dicts(prediction_path)

    pair_map = build_pair_map(pair_rows)
    validate_gold_rows(gold_rows, pair_map)
    gold_annotation_ids = {row["annotation_id"] for row in gold_rows}
    valid_pred_rows, invalid_pred_rows = validate_prediction_rows(
        prediction_rows,
        gold_annotation_ids=gold_annotation_ids,
        pair_map=pair_map,
    )

    exact_matches, _, _ = compute_exact_matches(gold_rows, valid_pred_rows)
    relaxed_matches = compute_relaxed_matches(gold_rows, valid_pred_rows, relaxed_threshold)
    per_pair_metrics = build_per_pair_metrics(
        gold_rows=gold_rows,
        valid_pred_rows=valid_pred_rows,
        invalid_pred_rows=invalid_pred_rows,
        exact_matches=exact_matches,
        relaxed_matches=relaxed_matches,
    )

    unmatched_gold = build_unmatched_gold_rows(gold_rows, relaxed_matches)
    overgenerated = build_overgenerated_pred_rows(valid_pred_rows, relaxed_matches)
    confusion_matrix = build_confusion_matrix(relaxed_matches)
    core_errors = build_core_error_report(gold_rows, valid_pred_rows, relaxed_matches)
    summary = compute_summary(
        gold_rows=gold_rows,
        valid_pred_rows=valid_pred_rows,
        invalid_pred_rows=invalid_pred_rows,
        exact_matches=exact_matches,
        relaxed_matches=relaxed_matches,
        per_pair_metrics=per_pair_metrics,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "evaluation_summary.json").write_text(
        json.dumps(
            {
                "run_name": run_name,
                "relaxed_threshold": relaxed_threshold,
                "summary": summary,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    write_summary_markdown(output_dir / "evaluation_summary.md", summary, run_name)

    write_csv(
        output_dir / "per_pair_metrics.csv",
        per_pair_metrics,
        fieldnames=list(per_pair_metrics[0].keys()),
    )
    write_xlsx(
        output_dir / "per_pair_metrics.xlsx",
        per_pair_metrics,
        fieldnames=list(per_pair_metrics[0].keys()),
        sheet_name="PerPairMetrics",
    )
    write_csv(
        output_dir / "matched_columns_exact.csv",
        build_match_rows(exact_matches),
        fieldnames=MATCH_FIELDNAMES,
    )
    write_csv(
        output_dir / "matched_columns_relaxed.csv",
        build_match_rows(relaxed_matches),
        fieldnames=MATCH_FIELDNAMES,
    )
    write_csv(
        output_dir / "unmatched_gold_columns.csv",
        unmatched_gold,
        fieldnames=list(unmatched_gold[0].keys()) if unmatched_gold else [*GOLD_REQUIRED_FIELDS, "unmatched_reason"],
    )
    write_csv(
        output_dir / "overgenerated_prediction_columns.csv",
        overgenerated,
        fieldnames=list(overgenerated[0].keys()) if overgenerated else [*PRED_REQUIRED_FIELDS, "overgenerated_reason"],
    )
    write_csv(
        output_dir / "invalid_predictions.csv",
        invalid_pred_rows,
        fieldnames=list(invalid_pred_rows[0].keys()) if invalid_pred_rows else [*PRED_REQUIRED_FIELDS, "invalid_reason"],
    )
    write_csv(
        output_dir / "relation_type_confusion_matrix.csv",
        confusion_matrix,
        fieldnames=list(confusion_matrix[0].keys()),
    )
    write_csv(
        output_dir / "core_column_error_report.csv",
        core_errors,
        fieldnames=list(core_errors[0].keys())
        if core_errors
        else [
            "error_type",
            "annotation_id",
            "pair_id",
            "gold_column_id",
            "pred_column_id",
            "gold_relation_type",
            "pred_relation_type",
            "gold_is_core_column",
            "pred_is_core_column",
            "details",
        ],
    )

    return {
        "summary": summary,
        "per_pair_metrics_path": output_dir / "per_pair_metrics.csv",
        "per_pair_metrics_xlsx_path": output_dir / "per_pair_metrics.xlsx",
        "output_dir": output_dir,
        "invalid_prediction_count": len(invalid_pred_rows),
        "valid_prediction_count": len(valid_pred_rows),
    }


def choose_alternative_relation_type(relation_type: str) -> str:
    for candidate in [
        "semantic_substitution",
        "slot_filling",
        "pragmatic_function",
        "coreference_or_demonstrative",
        "contrast",
        "repair",
    ]:
        if candidate != relation_type:
            return candidate
    return "other"


def find_relaxed_variant(span: str) -> str | None:
    span = span or ""
    if len(span) < 3:
        return None
    candidates = [
        span[:-1],
        span[1:],
        span[1:-1],
    ]
    for candidate in candidates:
        if candidate and candidate != span:
            return candidate
    return None


def find_overgenerated_pair(
    annotation_id: str,
    turn_a: str,
    turn_b: str,
    existing_pairs: set[Tuple[str, str]],
    gold_pair_rows: List[Dict[str, str]],
    relaxed_threshold: float,
) -> Tuple[str, str] | None:
    substrings_a: List[str] = []
    substrings_b: List[str] = []
    for min_len in (2, 3, 4):
        if len(turn_a) >= min_len:
            for start in range(0, len(turn_a) - min_len + 1):
                substr = turn_a[start : start + min_len]
                if substr.strip():
                    substrings_a.append(substr)
        if len(turn_b) >= min_len:
            for start in range(0, len(turn_b) - min_len + 1):
                substr = turn_b[start : start + min_len]
                if substr.strip():
                    substrings_b.append(substr)
    for span_a in substrings_a[:50]:
        for span_b in substrings_b[:50]:
            if (span_a, span_b) not in existing_pairs:
                would_relaxed_match = False
                for gold in gold_pair_rows:
                    overlap_a = span_overlap_ratio(gold["span_a"], span_a)
                    overlap_b = span_overlap_ratio(gold["span_b"], span_b)
                    if overlap_a >= relaxed_threshold and overlap_b >= relaxed_threshold:
                        would_relaxed_match = True
                        break
                if not would_relaxed_match:
                    return span_a, span_b
    return None


def make_toy_prediction(
    gold_path: Path,
    pair_list_path: Path,
    output_csv: Path,
) -> Dict[str, object]:
    gold_rows = read_csv_dicts(gold_path)
    pair_rows = read_csv_dicts(pair_list_path)
    pair_map = build_pair_map(pair_rows)
    validate_gold_rows(gold_rows, pair_map)

    toy_rows: List[Dict[str, str]] = []
    used_keys: set[str] = set()
    pred_counter = 1

    def next_pred_id() -> str:
        nonlocal pred_counter
        value = f"P{pred_counter:02d}"
        pred_counter += 1
        return value

    def pick_rows(
        limit: int,
        skip_used: bool = True,
        avoid_annotation_ids: set[str] | None = None,
    ) -> List[Dict[str, str]]:
        picked: List[Dict[str, str]] = []
        seen_annotation_ids: set[str] = set()
        blocked_annotation_ids = avoid_annotation_ids or set()
        for gold in gold_rows:
            if skip_used and gold_key(gold) in used_keys:
                continue
            if gold["annotation_id"] in blocked_annotation_ids:
                continue
            if gold["annotation_id"] in seen_annotation_ids:
                continue
            picked.append(gold)
            seen_annotation_ids.add(gold["annotation_id"])
            if len(picked) >= limit:
                break
        return picked

    # 1) exact correct rows
    exact_examples = pick_rows(limit=3)
    for gold in exact_examples:
        toy_rows.append(
            {
                "annotation_id": gold["annotation_id"],
                "pair_id": gold["pair_id"],
                "pred_column_id": next_pred_id(),
                "pred_span_a": gold["span_a"],
                "pred_span_b": gold["span_b"],
                "pred_relation_type": gold["relation_type"],
                "pred_relation_strength": gold["relation_strength"],
                "pred_alignment_direction": gold["alignment_direction"],
                "pred_is_core_column": gold["is_core_column"],
                "pred_supports_resonance": gold["supports_resonance"],
                "pred_confidence": "0.95",
                "generator_name": "toy_exact_subset",
                "generator_version": "v1",
                "notes": "exact_correct",
            }
        )
        used_keys.add(gold_key(gold))

    # 2) exact spans but wrong relation_type
    wrong_type_candidates = pick_rows(
        limit=2,
        avoid_annotation_ids={row["annotation_id"] for row in exact_examples},
    )
    for gold in wrong_type_candidates:
        toy_rows.append(
            {
                "annotation_id": gold["annotation_id"],
                "pair_id": gold["pair_id"],
                "pred_column_id": next_pred_id(),
                "pred_span_a": gold["span_a"],
                "pred_span_b": gold["span_b"],
                "pred_relation_type": choose_alternative_relation_type(gold["relation_type"]),
                "pred_relation_strength": gold["relation_strength"],
                "pred_alignment_direction": gold["alignment_direction"],
                "pred_is_core_column": gold["is_core_column"],
                "pred_supports_resonance": gold["supports_resonance"],
                "pred_confidence": "0.70",
                "generator_name": "toy_exact_subset",
                "generator_version": "v1",
                "notes": "exact_span_wrong_relation_type",
            }
        )
        used_keys.add(gold_key(gold))

    # 3) relaxed-only rows with partial overlap
    relaxed_candidates = [row for row in gold_rows if gold_key(row) not in used_keys]
    relaxed_added = 0
    relaxed_seen_pairs: set[str] = set()
    for gold in relaxed_candidates:
        if gold["annotation_id"] in relaxed_seen_pairs:
            continue
        relaxed_span_a = find_relaxed_variant(gold["span_a"])
        relaxed_span_b = find_relaxed_variant(gold["span_b"])
        if not relaxed_span_a or not relaxed_span_b:
            continue
        pair = pair_map[gold["annotation_id"]]
        if relaxed_span_a not in pair["turn_a"] or relaxed_span_b not in pair["turn_b"]:
            continue
        toy_rows.append(
            {
                "annotation_id": gold["annotation_id"],
                "pair_id": gold["pair_id"],
                "pred_column_id": next_pred_id(),
                "pred_span_a": relaxed_span_a,
                "pred_span_b": relaxed_span_b,
                "pred_relation_type": gold["relation_type"],
                "pred_relation_strength": gold["relation_strength"],
                "pred_alignment_direction": gold["alignment_direction"],
                "pred_is_core_column": gold["is_core_column"],
                "pred_supports_resonance": gold["supports_resonance"],
                "pred_confidence": "0.55",
                "generator_name": "toy_exact_subset",
                "generator_version": "v1",
                "notes": "relaxed_only_partial_overlap",
            }
        )
        used_keys.add(gold_key(gold))
        relaxed_seen_pairs.add(gold["annotation_id"])
        relaxed_added += 1
        if relaxed_added >= 2:
            break

    # 4) overgenerated valid rows
    overgenerated_added = 0
    by_pair_existing: Dict[str, set[Tuple[str, str]]] = defaultdict(set)
    for row in gold_rows:
        by_pair_existing[row["annotation_id"]].add((row["span_a"], row["span_b"]))
    for annotation_id, pair in pair_map.items():
        if overgenerated_added >= 2:
            break
        gold_pair_rows = [row for row in gold_rows if row["annotation_id"] == annotation_id]
        candidate = find_overgenerated_pair(
            annotation_id=annotation_id,
            turn_a=pair["turn_a"],
            turn_b=pair["turn_b"],
            existing_pairs=by_pair_existing[annotation_id],
            gold_pair_rows=gold_pair_rows,
            relaxed_threshold=0.5,
        )
        if not candidate:
            continue
        span_a, span_b = candidate
        by_pair_existing[annotation_id].add((span_a, span_b))
        toy_rows.append(
            {
                "annotation_id": annotation_id,
                "pair_id": pair["pair_id"],
                "pred_column_id": next_pred_id(),
                "pred_span_a": span_a,
                "pred_span_b": span_b,
                "pred_relation_type": "lexical_reproduction",
                "pred_relation_strength": "weak",
                "pred_alignment_direction": "A_to_B",
                "pred_is_core_column": "0",
                "pred_supports_resonance": "0",
                "pred_confidence": "0.22",
                "generator_name": "toy_exact_subset",
                "generator_version": "v1",
                "notes": "overgenerated_valid_prediction",
            }
        )
        overgenerated_added += 1

    # 5) invalid row for isolation test
    invalid_gold = gold_rows[0]
    toy_rows.append(
        {
            "annotation_id": invalid_gold["annotation_id"],
            "pair_id": invalid_gold["pair_id"],
            "pred_column_id": next_pred_id(),
            "pred_span_a": "不存在的A侧片段",
            "pred_span_b": invalid_gold["span_b"],
            "pred_relation_type": invalid_gold["relation_type"],
            "pred_relation_strength": invalid_gold["relation_strength"],
            "pred_alignment_direction": invalid_gold["alignment_direction"],
            "pred_is_core_column": invalid_gold["is_core_column"],
            "pred_supports_resonance": invalid_gold["supports_resonance"],
            "pred_confidence": "0.10",
            "generator_name": "toy_exact_subset",
            "generator_version": "v1",
            "notes": "invalid_prediction_for_validation",
        }
    )

    write_csv(output_csv, toy_rows, PRED_REQUIRED_FIELDS)
    write_xlsx(
        output_csv.with_suffix(".xlsx"),
        toy_rows,
        fieldnames=PRED_REQUIRED_FIELDS,
        sheet_name="ToyPrediction",
    )
    return {
        "output_csv": str(output_csv),
        "output_xlsx": str(output_csv.with_suffix(".xlsx")),
        "row_count": len(toy_rows),
        "exact_rows": len(exact_examples),
        "wrong_type_rows": len(wrong_type_candidates),
        "relaxed_rows": relaxed_added,
        "overgenerated_rows": overgenerated_added,
        "invalid_rows": 1,
    }


def write_implementation_report(
    summary_json_path: Path,
    output_path: Path,
) -> None:
    summary_data = json.loads(summary_json_path.read_text(encoding="utf-8"))
    summary = summary_data["summary"]
    lines = [
        "# diagraph evaluation runner v1 implementation report",
        "",
        "## Supported functionality",
        "",
        "- 读取并校验 `diagraph_gold_50_column_gold_v1_active.csv`。",
        "- 读取并校验符合 `diagraph_prediction_schema_v1.md` 的 prediction file。",
        "- 将不合法 prediction 隔离到 `invalid_predictions.csv`，不参与 metric。",
        "- 计算 exact matching、relaxed matching，以及 relation/core 相关指标。",
        "- 输出 `evaluation_summary.*`、`per_pair_metrics.csv/xlsx`、matched / unmatched / overgenerated / confusion / core error 报告。",
        "- 支持生成 toy prediction CSV/XLSX，用于离线链路测试。",
        "",
        "## exact / relaxed matching logic",
        "",
        "- exact matching：`gold span_a == pred span_a` 且 `gold span_b == pred span_b`。",
        "- relaxed matching：exact 之后，对剩余列按 A/B 两侧 span overlap 做保守 one-to-one 匹配。",
        "- overlap ratio 优先处理 exact 与包含关系；其余情形使用较短 span 上的最长公共子串覆盖率。",
        "- 默认 `relaxed_threshold = 0.5`。",
        "- 同一 gold column 最多匹配一个 pred，同一 pred column 最多匹配一个 gold；同分时优先更高 overlap。",
        "",
        "## toy run result",
        "",
        f"- gold columns: {summary['gold_column_count']}",
        f"- valid predictions: {summary['valid_prediction_count']}",
        f"- invalid predictions: {summary['invalid_prediction_count']}",
        f"- exact precision / recall / F1: {summary['exact_column_precision']} / {summary['exact_column_recall']} / {summary['exact_column_f1']}",
        f"- relaxed precision / recall / F1: {summary['relaxed_column_precision']} / {summary['relaxed_column_recall']} / {summary['relaxed_column_f1']}",
        f"- relation_type accuracy on exact matches: {summary['relation_type_accuracy_on_exact_matches']}",
        f"- relation_type accuracy on relaxed matches: {summary['relation_type_accuracy_on_relaxed_matches']}",
        f"- core precision / recall: {summary['core_column_precision']} / {summary['core_column_recall']}",
        f"- overgeneration rate: {summary['overgeneration_rate']}",
        f"- missing column rate: {summary['missing_column_rate']}",
        "",
        "## Current limitations",
        "",
        "- 当前 relaxed matching 是保守字符级规则，不处理语义等价、释义或 paraphrase。",
        "- 当前未单独输出 alignment direction accuracy；direction 信息会保留在 matched 报表中供后续扩展。",
        "- 当前 confusion matrix 基于 relaxed matched columns。",
        "- 当前 core metrics 也基于 relaxed matched columns，更适合结构恢复评估，不等同于严格 exact core scoring。",
        "- toy prediction 只用于链路测试，不代表真实 baseline 质量。",
        "",
        "## How to connect future rule baseline predictions",
        "",
        "1. 让 rule baseline 输出符合 `diagraph_prediction_schema_v1.md` 的 prediction file。",
        "2. 将 prediction file 传入 evaluator。",
        "3. 读取 `evaluation_summary.*`、`per_pair_metrics.csv/xlsx`、`unmatched_gold_columns.csv`、`overgenerated_prediction_columns.csv` 做 error analysis。",
        "",
        "## Scope note",
        "",
        "- 本轮未训练或运行 BERT。",
        "- 本轮未修改 `diagraph_gold_50_column_gold_v1`。",
        "- 本轮未接网站。",
        "- 本轮未读取或写入正式数据库。",
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate diagraph generation against gold_v1.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    toy_parser = subparsers.add_parser("make-toy", help="Generate a toy prediction file.")
    toy_parser.add_argument("--gold-active", type=Path, default=DEFAULT_GOLD_ACTIVE)
    toy_parser.add_argument("--pair-list", type=Path, default=PAIR_LIST_PATH)
    toy_parser.add_argument("--output-csv", type=Path, default=DEFAULT_TOY_PREDICTION)

    eval_parser = subparsers.add_parser("evaluate", help="Run the offline evaluator.")
    eval_parser.add_argument("--gold-active", type=Path, default=DEFAULT_GOLD_ACTIVE)
    eval_parser.add_argument("--pair-list", type=Path, default=PAIR_LIST_PATH)
    eval_parser.add_argument("--prediction", type=Path, required=True)
    eval_parser.add_argument("--output-dir", type=Path, required=True)
    eval_parser.add_argument("--relaxed-threshold", type=float, default=0.5)
    eval_parser.add_argument("--run-name", default="diagraph_eval_run")

    report_parser = subparsers.add_parser(
        "implementation-report",
        help="Write an implementation report from a completed evaluation summary.",
    )
    report_parser.add_argument("--summary-json", type=Path, required=True)
    report_parser.add_argument("--output-md", type=Path, default=DEFAULT_IMPLEMENTATION_REPORT)

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.command == "make-toy":
        result = make_toy_prediction(
            gold_path=args.gold_active,
            pair_list_path=args.pair_list,
            output_csv=args.output_csv,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if args.command == "evaluate":
        result = evaluate_predictions(
            gold_path=args.gold_active,
            pair_list_path=args.pair_list,
            prediction_path=args.prediction,
            output_dir=args.output_dir,
            relaxed_threshold=args.relaxed_threshold,
            run_name=args.run_name,
        )
        print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
        return

    if args.command == "implementation-report":
        write_implementation_report(
            summary_json_path=args.summary_json,
            output_path=args.output_md,
        )
        print(json.dumps({"output_md": str(args.output_md)}, ensure_ascii=False, indent=2))
        return

    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    main()
