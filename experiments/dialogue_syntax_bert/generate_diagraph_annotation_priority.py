"""Generate annotation priority tiers and pilot10 suggestions for diagraph_gold_50."""

from __future__ import annotations

import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from io_utils import artifact_path, ensure_can_write, read_csv, write_csv, write_text


DIFFICULTY_ORDER = {"easy": 0, "medium": 1, "hard": 2}

DEMONSTRATIVE_PATTERN = re.compile(
    r"(这|那|此|其|之|他|她|它|这样|那样|这么|那么|这个|那个|这里|那里|如此|前者|后者)"
)
WH_PATTERN = re.compile(r"(什么|怎么|怎样|为何|为啥|谁|哪|哪里|哪个|多少|几|几时|何如|何以|何故|奚|孰)")

PRIORITY_FIELDS = [
    "annotation_id",
    "pair_id",
    "source",
    "dataset_name",
    "turn_a",
    "turn_b",
    "difficulty_level",
    "priority_rank",
    "expected_column_count",
    "dominant_relation_types",
    "why_this_difficulty",
    "annotation_warning",
    "suggested_first_pass",
]

PILOT10_IDS = [
    "F300V1-0050",
    "F300V1-0137",
    "F300V1-0020",
    "F300V1-0244",
    "F300V1-0170",
    "F300V1-0023",
    "F300V1-0196",
    "F300V1-0254",
    "F300V1-0127",
    "F300V1-0211",
]

MANUAL_DIFFICULTY_OVERRIDES = {
    "F300V1-0050": "easy",
    "F300V1-0137": "easy",
    "F300V1-0020": "easy",
    "F300V1-0244": "easy",
    "F300V1-0170": "easy",
    "F300V1-0023": "medium",
    "F300V1-0196": "medium",
    "F300V1-0254": "medium",
    "F300V1-0127": "hard",
    "F300V1-0211": "hard",
    "F300V1-0002": "medium",
    "F300V1-0055": "medium",
    "F300V1-0092": "hard",
    "F300V1-0106": "hard",
    "F300V1-0111": "hard",
    "F300V1-0117": "hard",
    "F300V1-0128": "hard",
    "F300V1-0150": "hard",
    "F300V1-0159": "hard",
    "F300V1-0187": "medium",
    "F300V1-0204": "medium",
    "F300V1-0205": "hard",
    "F300V1-0214": "hard",
    "F300V1-0220": "hard",
    "F300V1-0224": "hard",
    "F300V1-0250": "medium",
    "F300V1-0265": "medium",
    "F300V1-0287": "medium",
    "F300V1-0298": "medium",
    "F300V1-0299": "medium",
    "F300V1-0300": "medium",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    output_dir = artifact_path("formal_300_v1", "diagraph_gold_50")
    parser.add_argument("--output-dir", default=str(output_dir))
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def bool_text(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def has_demonstrative(text: str) -> bool:
    return bool(DEMONSTRATIVE_PATTERN.search(text))


def has_slot_filling(turn_a: str, turn_b: str) -> bool:
    if "？" in turn_a or "?" in turn_a or WH_PATTERN.search(turn_a):
        return "？" not in turn_b and "?" not in turn_b
    return False


def has_short_answer(turn_a: str, turn_b: str) -> bool:
    return min(len(turn_a.strip()), len(turn_b.strip())) <= 8


def semantic_substitution(row: dict[str, str]) -> bool:
    return row["label_selective_reuse"] == "1" and (
        row["label_reproduction"] == "0"
        or row["rule_any_positive"] == "0"
        or row["sample_stratum"] in {"potential_false_negative", "rule_negative_random"}
    )


def infer_relation_types(row: dict[str, str]) -> list[str]:
    relation_types: list[str] = []
    full_text = " ".join(
        [row["turn_a"], row["turn_b"], row["evidence_span_a"], row["evidence_span_b"]]
    )
    if row["label_analogy_candidate"] == "1":
        relation_types.append("analogy")
    if row["label_repair"] == "1":
        relation_types.append("repair")
    if row["label_contrast"] == "1":
        relation_types.append("contrast")
    if row["label_reproduction"] == "1":
        relation_types.append("lexical_reproduction")
    if row["label_parallelism"] == "1":
        relation_types.append("syntactic_parallelism")
    if row["label_selective_reuse"] == "1":
        relation_types.append("selective_reuse")
    if semantic_substitution(row):
        relation_types.append("semantic_substitution")
    if has_demonstrative(full_text):
        relation_types.append("demonstrative/reference")
    if has_slot_filling(row["turn_a"], row["turn_b"]):
        relation_types.append("slot_filling")
    if has_short_answer(row["turn_a"], row["turn_b"]):
        relation_types.append("short_answer")
    if not relation_types:
        relation_types.append("pragmatic_function")
    deduped: list[str] = []
    for item in relation_types:
        if item not in deduped:
            deduped.append(item)
    return deduped[:4]


def estimate_column_count(row: dict[str, str]) -> int:
    count = 1
    count += int(row["label_reproduction"] == "1")
    count += int(row["label_parallelism"] == "1")
    count += int(row["label_contrast"] == "1")
    count += int(row["label_repair"] == "1")
    count += int(row["label_analogy_candidate"] == "1")
    count += int(has_slot_filling(row["turn_a"], row["turn_b"]))
    count += int(has_demonstrative(" ".join([row["turn_a"], row["turn_b"], row["evidence_span_a"], row["evidence_span_b"]])) and row["label_reproduction"] == "0")
    count += int(semantic_substitution(row) and row["label_reproduction"] == "0")
    pipe_count = row["evidence_span_a"].count("|||") + row["evidence_span_b"].count("|||")
    if pipe_count >= 3:
        count += 1
    if max(len(row["turn_a"]), len(row["turn_b"])) >= 80 and count >= 4:
        count += 1
    return min(count, 7)


def default_difficulty(row: dict[str, str], dominant_types: list[str], expected_columns: int) -> str:
    full_text = " ".join([row["turn_a"], row["turn_b"], row["evidence_span_a"], row["evidence_span_b"]])
    lexical = row["label_reproduction"] == "1"
    parallel = row["label_parallelism"] == "1"
    analogy = row["label_analogy_candidate"] == "1"
    repair = row["label_repair"] == "1"
    contrast = row["label_contrast"] == "1"
    demo = has_demonstrative(full_text)
    short_answer = has_short_answer(row["turn_a"], row["turn_b"])
    semantic = semantic_substitution(row)
    slot = has_slot_filling(row["turn_a"], row["turn_b"])
    low_surface = row["rule_any_positive"] == "0" and not lexical and not parallel
    long_turn = max(len(row["turn_a"]), len(row["turn_b"])) >= 70
    evidence_segments = row["evidence_span_a"].count("|||") + row["evidence_span_b"].count("|||")

    if analogy or (demo and (short_answer or semantic)) or (short_answer and low_surface) or (semantic and low_surface) or (long_turn and low_surface):
        return "hard"
    if (lexical or parallel) and expected_columns <= 3 and evidence_segments <= 2 and not analogy and not (demo and not lexical) and not (short_answer and low_surface):
        return "easy"
    if slot or contrast or repair or expected_columns <= 5:
        return "medium"
    return "hard"


def explain_difficulty(row: dict[str, str], difficulty: str, dominant_types: list[str], expected_columns: int) -> str:
    dominant = " / ".join(dominant_types)
    if difficulty == "easy":
        return (
            f"表面映射较清楚，主导关系以 {dominant} 为主；evidence span 边界较稳，"
            f"预计纵栏约 {expected_columns} 个，适合先做示范性标注。"
        )
    if difficulty == "medium":
        return (
            f"存在 {dominant} 这类需要判断主次关系的映射；通常能找到稳定纵栏，"
            f"但 relation_type 还需人工区分，预计纵栏约 {expected_columns} 个。"
        )
    return (
        f"该样本更依赖 {dominant} 这类隐性映射或上下文推断；"
        f"表面重合较低，容易和单纯话题相关混淆，预计纵栏约 {expected_columns} 个。"
    )


def annotation_warning(row: dict[str, str], dominant_types: list[str], expected_columns: int) -> str:
    full_text = " ".join([row["turn_a"], row["turn_b"], row["evidence_span_a"], row["evidence_span_b"]])
    warnings: list[str] = []
    if "analogy" in dominant_types:
        warnings.append("先确认是否真存在类比映射，不要只因评价或比喻出现就多画栏")
    if "demonstrative/reference" in dominant_types:
        warnings.append("指代词需先确认回指对象，不能自动替换成你理解的实体")
    if "short_answer" in dominant_types:
        warnings.append("短答高度依赖上句，先判断是否形成稳定纵栏而非普通接话")
    if "semantic_substitution" in dominant_types:
        warnings.append("注意区分语义替代与纯话题承接")
    if "repair" in dominant_types:
        warnings.append("区分修正/否定重构与单纯反驳")
    if max(len(row["turn_a"]), len(row["turn_b"])) >= 90:
        warnings.append("优先截取主干成分，避免把整段长句整块贴入 span")
    if expected_columns >= 6:
        warnings.append("该样本可能超过 5 行模板，必要时需手动扩展")
    if not warnings and has_demonstrative(full_text):
        warnings.append("注意这里可能存在隐性指代，不要直接按同主题处理")
    return "；".join(warnings[:3]) if warnings else "先从 evidence span 入手，避免扩大到整段话题范围"


def clarity_score(row: dict[str, str], difficulty: str, expected_columns: int) -> tuple[float, float, float]:
    lexical = row["label_reproduction"] == "1"
    parallel = row["label_parallelism"] == "1"
    repair = row["label_repair"] == "1"
    contrast = row["label_contrast"] == "1"
    analogy = row["label_analogy_candidate"] == "1"
    score = 0.0
    score += 3.0 if lexical else 0.0
    score += 2.0 if parallel else 0.0
    score += 1.2 if row["label_selective_reuse"] == "1" else 0.0
    score += 1.0 if repair else 0.0
    score += 0.8 if contrast else 0.0
    score -= 1.5 if analogy else 0.0
    score -= 1.0 if has_short_answer(row["turn_a"], row["turn_b"]) and not lexical else 0.0
    score -= 1.0 if semantic_substitution(row) and not lexical else 0.0
    score -= 0.6 * max(0, expected_columns - 2)
    difficulty_bonus = {"easy": 3.0, "medium": 1.5, "hard": 0.0}[difficulty]
    surface_bonus = 0.3 if row["rule_any_positive"] == "1" else 0.0
    return (difficulty_bonus + score + surface_bonus, -expected_columns, float(row["bert_prob"] or 0.0))


def write_workbook(path: Path, rows: list[dict[str, str]], fieldnames: list[str], *, overwrite: bool) -> None:
    output_path = ensure_can_write(path, overwrite=overwrite)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "priority"
    fill = PatternFill(fill_type="solid", start_color="F7E6D5", end_color="F7E6D5")
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    widths = {
        "annotation_id": 16,
        "pair_id": 12,
        "source": 12,
        "dataset_name": 18,
        "turn_a": 42,
        "turn_b": 42,
        "difficulty_level": 12,
        "priority_rank": 12,
        "expected_column_count": 18,
        "dominant_relation_types": 34,
        "why_this_difficulty": 44,
        "annotation_warning": 44,
        "suggested_first_pass": 14,
    }
    for idx, field in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=idx, value=field)
        cell.fill = fill
        cell.font = bold
        cell.alignment = wrap
        sheet.column_dimensions[cell.column_letter].width = widths.get(field, 18)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = wrap
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def build_priority_rows(pair_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    enriched_rows: list[dict[str, str]] = []
    for row in pair_rows:
        dominant_types = infer_relation_types(row)
        expected_columns = estimate_column_count(row)
        difficulty = MANUAL_DIFFICULTY_OVERRIDES.get(
            row["annotation_id"],
            default_difficulty(row, dominant_types, expected_columns),
        )
        enriched_rows.append(
            {
                "annotation_id": row["annotation_id"],
                "pair_id": row["pair_id"],
                "source": row["source"],
                "dataset_name": row["dataset_name"],
                "turn_a": row["turn_a"],
                "turn_b": row["turn_b"],
                "difficulty_level": difficulty,
                "priority_rank": "",
                "expected_column_count": str(expected_columns),
                "dominant_relation_types": " | ".join(dominant_types),
                "why_this_difficulty": explain_difficulty(row, difficulty, dominant_types, expected_columns),
                "annotation_warning": annotation_warning(row, dominant_types, expected_columns),
                "suggested_first_pass": "1" if row["annotation_id"] in PILOT10_IDS else "0",
            }
        )

    def sort_key(item: dict[str, str]) -> tuple[int, int, tuple[float, float, float], str]:
        original = next(row for row in pair_rows if row["annotation_id"] == item["annotation_id"])
        clarity = clarity_score(original, item["difficulty_level"], int(item["expected_column_count"]))
        return (
            DIFFICULTY_ORDER[item["difficulty_level"]],
            -int(item["suggested_first_pass"]),
            tuple(-x for x in clarity),
            item["annotation_id"],
        )

    enriched_rows.sort(key=sort_key)
    for idx, item in enumerate(enriched_rows, start=1):
        item["priority_rank"] = str(idx)
    return enriched_rows


def build_priority_report(
    priority_rows: list[dict[str, str]],
    pilot_rows: list[dict[str, str]],
) -> str:
    counts = Counter(row["difficulty_level"] for row in priority_rows)
    example_ids = ["F300V1-0050", "F300V1-0137", "F300V1-0244", "F300V1-0196"]
    wait_ids = ["F300V1-0127", "F300V1-0092", "F300V1-0211", "F300V1-0220", "F300V1-0205", "F300V1-0224"]
    extension_rows = [row for row in priority_rows if int(row["expected_column_count"]) >= 6]
    pilot_lines = [
        f"- `{row['annotation_id']}` / {row['source']} / {row['dataset_name']}: {row['dominant_relation_types']}"
        for row in pilot_rows
    ]
    example_lines = [
        f"- `{row['annotation_id']}`: {row['dominant_relation_types']} ({row['why_this_difficulty']})"
        for row in priority_rows
        if row["annotation_id"] in example_ids
    ]
    wait_lines = [
        f"- `{row['annotation_id']}`: {row['dominant_relation_types']}；{row['annotation_warning']}"
        for row in priority_rows
        if row["annotation_id"] in wait_ids
    ]
    extension_lines = [
        f"- `{row['annotation_id']}`: expected_column_count={row['expected_column_count']}，{row['dominant_relation_types']}"
        for row in extension_rows
    ]
    return f"""# diagraph_gold_50 priority report

## 1. 难度分布

- easy: {counts['easy']}
- medium: {counts['medium']}
- hard: {counts['hard']}

## 2. pilot10 为什么这样选

pilot10 按 `easy 5 / medium 3 / hard 2` 取样，目的是让第一轮人工标注同时覆盖：

- lexical reproduction
- syntactic parallelism
- selective reuse
- contrast
- slot filling
- demonstrative/reference
- short answer
- analogy

同时尽量避免全部集中在同一 dataset，并保留文本对话、日常对话、影视对白等不同来源的对照感。

pilot10 列表：

{chr(10).join(pilot_lines)}

## 3. 最适合作为标注示例的样本

以下样本建议在正式开标前先一起过一遍，作为“什么叫清晰纵栏”的示例：

{chr(10).join(example_lines)}

## 4. 建议等 guide 修订后再标的样本

以下样本更依赖隐性回指、短答语义补足、类比或低表面重合，建议先用 pilot10 校正 guide，再处理这些 harder cases：

{chr(10).join(wait_lines)}

## 5. 哪些样本可能需要超过 5 行

如果每个 pair 只预留 5 行，以下样本最可能需要扩展：

{chr(10).join(extension_lines) if extension_lines else '- 当前没有明显超过 5 行风险的样本。'}

## 6. 下一步建议

建议流程：

1. 先标 pilot10；
2. 根据 pilot10 的分歧修订 guide；
3. 再标剩余 40 条；
4. 最后回看是否需要对 `expected_column_count >= 6` 的样本单独加行或拆分复核。
"""


def validate_inputs(pair_rows: list[dict[str, str]], column_rows: list[dict[str, str]]) -> None:
    if len(pair_rows) != 50:
        raise SystemExit(f"Expected 50 pair rows, got {len(pair_rows)}")
    if len(column_rows) < 50:
        raise SystemExit("Column template looks incomplete.")
    pair_ids = {row["annotation_id"] for row in pair_rows}
    column_ids = {row["annotation_id"] for row in column_rows}
    if not pair_ids.issubset(column_ids):
        missing = sorted(pair_ids - column_ids)
        raise SystemExit(f"Column template missing pair ids: {missing}")
    if any(pilot_id not in pair_ids for pilot_id in PILOT10_IDS):
        raise SystemExit("Pilot10 ids are not all present in diagraph_gold_50 pair list.")


def validate_outputs(priority_rows: list[dict[str, str]], pilot_rows: list[dict[str, str]]) -> None:
    if len(priority_rows) != 50:
        raise SystemExit(f"Priority file must contain 50 rows, got {len(priority_rows)}")
    if len(pilot_rows) != 10:
        raise SystemExit(f"Pilot10 file must contain 10 rows, got {len(pilot_rows)}")
    pilot_counts = Counter(row["difficulty_level"] for row in pilot_rows)
    if pilot_counts["easy"] != 5 or pilot_counts["medium"] != 3 or pilot_counts["hard"] != 2:
        raise SystemExit(f"Pilot10 tier quota mismatch: {pilot_counts}")
    if len({row["annotation_id"] for row in pilot_rows}) != 10:
        raise SystemExit("Pilot10 contains duplicate annotation_id values.")


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    expected_dir = artifact_path("formal_300_v1", "diagraph_gold_50").resolve()
    if output_dir.resolve() != expected_dir:
        raise SystemExit(f"Outputs must stay under {expected_dir}")

    pair_csv = output_dir / "diagraph_gold_50_pair_list.csv"
    column_csv = output_dir / "diagraph_gold_50_column_annotation_template.csv"
    guide_md = output_dir / "diagraph_gold_50_annotation_guide.md"

    pair_rows = read_csv(pair_csv)
    column_rows = read_csv(column_csv)
    if not guide_md.exists():
        raise SystemExit(f"Missing annotation guide: {guide_md}")
    validate_inputs(pair_rows, column_rows)

    priority_rows = build_priority_rows(pair_rows)
    pilot_rows = [row for row in priority_rows if row["suggested_first_pass"] == "1"]
    validate_outputs(priority_rows, pilot_rows)
    report_text = build_priority_report(priority_rows, pilot_rows)

    priority_csv = output_dir / "diagraph_gold_50_annotation_priority.csv"
    priority_xlsx = output_dir / "diagraph_gold_50_annotation_priority.xlsx"
    pilot_csv = output_dir / "diagraph_gold_50_pilot10_list.csv"
    pilot_xlsx = output_dir / "diagraph_gold_50_pilot10_list.xlsx"
    report_md = output_dir / "diagraph_gold_50_priority_report.md"

    write_csv(priority_csv, priority_rows, PRIORITY_FIELDS, overwrite=args.overwrite)
    write_workbook(priority_xlsx, priority_rows, PRIORITY_FIELDS, overwrite=args.overwrite)
    write_csv(pilot_csv, pilot_rows, PRIORITY_FIELDS, overwrite=args.overwrite)
    write_workbook(pilot_xlsx, pilot_rows, PRIORITY_FIELDS, overwrite=args.overwrite)
    write_text(report_md, report_text, overwrite=args.overwrite)

    counts = Counter(row["difficulty_level"] for row in priority_rows)
    print("priority generation complete")
    print("difficulty distribution:", dict(counts))
    print("pilot10 ids:", ",".join(row["annotation_id"] for row in pilot_rows))


if __name__ == "__main__":
    main()
