"""Generate the diagraph_gold_50 sampling packet and annotation templates.

This script stays fully offline:

- reads only frozen CSV/XLSX experiment artifacts
- does not touch corpus.db
- does not train or run BERT
- reconstructs shadow-only bert_prob / hybrid_pred from existing prediction CSVs
"""

from __future__ import annotations

import argparse
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from io_utils import artifact_path, ensure_can_write, read_csv, write_csv, write_text


TARGET_TYPES = [
    "lexical_reproduction",
    "syntactic_parallelism",
    "selective_reuse",
    "repair",
    "contrast",
    "analogy_candidate",
    "demonstrative_reference",
    "slot_filling",
    "short_answer",
    "semantic_substitution",
]

TYPE_TARGETS = {
    "lexical_reproduction": 14,
    "syntactic_parallelism": 8,
    "selective_reuse": 24,
    "repair": 6,
    "contrast": 10,
    "analogy_candidate": 5,
    "demonstrative_reference": 8,
    "slot_filling": 8,
    "short_answer": 8,
    "semantic_substitution": 10,
}

TYPE_WEIGHTS = {
    "lexical_reproduction": 1.0,
    "syntactic_parallelism": 1.6,
    "selective_reuse": 0.4,
    "repair": 2.2,
    "contrast": 1.4,
    "analogy_candidate": 2.6,
    "demonstrative_reference": 1.5,
    "slot_filling": 1.4,
    "short_answer": 1.2,
    "semantic_substitution": 1.7,
}

RELATION_TYPES = [
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
]
RELATION_STRENGTHS = ["strong", "medium", "weak"]
ALIGNMENT_DIRECTIONS = ["A_to_B", "B_to_A", "mutual"]
TERNARY_VALUES = ["1", "0", "?"]

HYBRID_STRATEGY_NAME = "rule_priority_with_bert_recall"
HYBRID_THRESHOLD = 0.64037926197052
TEMPLATE_ROWS_PER_PAIR = 5

CURATED_PRIORITY_IDS = [
    "F300V1-0002",
    "F300V1-0008",
    "F300V1-0013",
    "F300V1-0017",
    "F300V1-0020",
    "F300V1-0023",
    "F300V1-0024",
    "F300V1-0033",
    "F300V1-0043",
    "F300V1-0046",
    "F300V1-0050",
    "F300V1-0052",
    "F300V1-0053",
    "F300V1-0055",
    "F300V1-0081",
    "F300V1-0092",
    "F300V1-0127",
    "F300V1-0150",
    "F300V1-0187",
    "F300V1-0220",
    "F300V1-0244",
    "F300V1-0254",
    "F300V1-0265",
    "F300V1-0299",
]

PAIR_LIST_FIELDS = [
    "annotation_id",
    "pair_id",
    "source",
    "dataset_name",
    "sample_stratum",
    "turn_a",
    "turn_b",
    "resonance_present",
    "label_reproduction",
    "label_parallelism",
    "label_selective_reuse",
    "label_repair",
    "label_contrast",
    "label_analogy_candidate",
    "evidence_span_a",
    "evidence_span_b",
    "annotator_note",
    "rule_any_positive",
    "bert_prob",
    "hybrid_pred",
]

COLUMN_TEMPLATE_FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
    "notes",
]

DEMONSTRATIVE_PATTERN = re.compile(
    r"(这|那|此|其|之|他|她|它|这样|那样|这么|那么|这个|那个|这里|那里|如此|如此一来|前者|后者)"
)
WH_PATTERN = re.compile(r"(什么|怎么|怎样|为何|为啥|谁|哪|哪里|哪个|多少|几|几时|何如|何以|何故|奚|孰)")
CONTRAST_PATTERN = re.compile(r"(不|没|无|却|反而|还是|不是|而是|不过|岂|未若)")


@dataclass(frozen=True)
class Candidate:
    annotation_id: str
    pair_id: str
    source: str
    dataset_name: str
    sample_stratum: str
    turn_a: str
    turn_b: str
    resonance_present: str
    label_reproduction: str
    label_parallelism: str
    label_selective_reuse: str
    label_repair: str
    label_contrast: str
    label_analogy_candidate: str
    evidence_span_a: str
    evidence_span_b: str
    annotator_note: str
    rule_any_positive: str
    bert_prob: str
    hybrid_pred: str
    split: str
    target_types: tuple[str, ...]

    def to_pair_list_row(self) -> dict[str, str]:
        return {
            "annotation_id": self.annotation_id,
            "pair_id": self.pair_id,
            "source": self.source,
            "dataset_name": self.dataset_name,
            "sample_stratum": self.sample_stratum,
            "turn_a": self.turn_a,
            "turn_b": self.turn_b,
            "resonance_present": self.resonance_present,
            "label_reproduction": self.label_reproduction,
            "label_parallelism": self.label_parallelism,
            "label_selective_reuse": self.label_selective_reuse,
            "label_repair": self.label_repair,
            "label_contrast": self.label_contrast,
            "label_analogy_candidate": self.label_analogy_candidate,
            "evidence_span_a": self.evidence_span_a,
            "evidence_span_b": self.evidence_span_b,
            "annotator_note": self.annotator_note,
            "rule_any_positive": self.rule_any_positive,
            "bert_prob": self.bert_prob,
            "hybrid_pred": self.hybrid_pred,
        }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    base = artifact_path("formal_300_v1")
    parser.add_argument(
        "--output-dir",
        default=str(base / "diagraph_gold_50"),
        help="Output directory under experiments/dialogue_syntax_bert/artifacts/",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow overwriting generated artifacts if they already exist.",
    )
    return parser.parse_args()


def bool_text(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def load_split_map(base: Path) -> dict[str, str]:
    split_map: dict[str, str] = {}
    baselines_dir = base / "baselines"
    for split in ("train", "dev", "test"):
        for row in read_csv(baselines_dir / f"gold_v1_binary_{split}.csv"):
            split_map[row["annotation_id"]] = split
    return split_map


def load_mean_bert_probabilities(multiseed_dir: Path) -> dict[str, float]:
    prob_values: dict[str, list[float]] = defaultdict(list)
    for seed_dir in sorted(multiseed_dir.glob("seed_*")):
        for split_name in ("dev", "test"):
            path = seed_dir / f"{split_name}_predictions.csv"
            if not path.exists():
                continue
            for row in read_csv(path):
                annotation_id = row["annotation_id"]
                prob_values[annotation_id].append(float(row["prob_yes"]))
    return {
        annotation_id: statistics.mean(values)
        for annotation_id, values in prob_values.items()
        if values
    }


def load_sources(base: Path) -> list[Candidate]:
    gold_binary_rows = {row["annotation_id"]: row for row in read_csv(base / "formal_300_v1_gold_v1_binary.csv")}
    gold_full_rows = {row["annotation_id"]: row for row in read_csv(base / "formal_300_v1_gold_v1.csv")}
    eval_rows = {row["annotation_id"]: row for row in read_csv(base / "formal_300_v1_evaluation_key.csv")}
    split_map = load_split_map(base)
    mean_probs = load_mean_bert_probabilities(base / "bert_shadow_v3_multiseed")

    candidates: list[Candidate] = []
    for annotation_id, row in gold_binary_rows.items():
        if row["resonance_present"] != "yes":
            continue
        full = gold_full_rows[annotation_id]
        eval_row = eval_rows.get(annotation_id, {})
        rule_any_positive = eval_row.get("rule_any_positive", "")
        mean_prob = mean_probs.get(annotation_id)
        bert_prob = f"{mean_prob:.6f}" if mean_prob is not None else ""
        hybrid_pred = ""
        if mean_prob is not None:
            hybrid_pred = "yes" if (bool_text(rule_any_positive) or mean_prob >= HYBRID_THRESHOLD) else "no"
        candidate = Candidate(
            annotation_id=annotation_id,
            pair_id=row["pair_id"],
            source=row["source"],
            dataset_name=row["dataset_name"],
            sample_stratum=row["sample_stratum"],
            turn_a=row["turn_a"],
            turn_b=row["turn_b"],
            resonance_present=row["resonance_present"],
            label_reproduction=row["label_reproduction"],
            label_parallelism=row["label_parallelism"],
            label_selective_reuse=row["label_selective_reuse"],
            label_repair=row["label_repair"],
            label_contrast=row["label_contrast"],
            label_analogy_candidate=row["label_analogy_candidate"],
            evidence_span_a=row["evidence_span_a"],
            evidence_span_b=row["evidence_span_b"],
            annotator_note=full.get("annotator_note", ""),
            rule_any_positive=rule_any_positive,
            bert_prob=bert_prob,
            hybrid_pred=hybrid_pred,
            split=split_map.get(annotation_id, ""),
            target_types=tuple(infer_target_types(row, eval_row)),
        )
        candidates.append(candidate)
    return candidates


def infer_target_types(row: dict[str, str], eval_row: dict[str, str]) -> list[str]:
    types: list[str] = []
    text = " ".join(
        [
            row.get("turn_a", ""),
            row.get("turn_b", ""),
            row.get("evidence_span_a", ""),
            row.get("evidence_span_b", ""),
        ]
    )
    turn_a = row.get("turn_a", "")
    turn_b = row.get("turn_b", "")
    if row.get("label_reproduction") == "1" or bool_text(eval_row.get("has_lexical_echo")):
        types.append("lexical_reproduction")
    if row.get("label_parallelism") == "1" or bool_text(eval_row.get("has_pattern_reuse")):
        types.append("syntactic_parallelism")
    if row.get("label_selective_reuse") == "1":
        types.append("selective_reuse")
    if row.get("label_repair") == "1" or bool_text(eval_row.get("has_repair_repetition")):
        types.append("repair")
    if row.get("label_contrast") == "1" or bool_text(eval_row.get("has_negation_turn")) or CONTRAST_PATTERN.search(text):
        types.append("contrast")
    if row.get("label_analogy_candidate") == "1":
        types.append("analogy_candidate")
    if DEMONSTRATIVE_PATTERN.search(text):
        types.append("demonstrative_reference")
    if (
        bool_text(eval_row.get("has_question_response"))
        or "？" in turn_a
        or "?" in turn_a
        or WH_PATTERN.search(turn_a)
    ) and ("？" not in turn_b and "?" not in turn_b):
        types.append("slot_filling")
    if min(len(turn_a.strip()), len(turn_b.strip())) <= 6:
        types.append("short_answer")
    if row.get("label_selective_reuse") == "1" and (
        row.get("label_reproduction") == "0"
        or not bool_text(eval_row.get("rule_any_positive"))
        or row.get("sample_stratum") in {"potential_false_negative", "rule_negative_random"}
    ):
        types.append("semantic_substitution")
    return sorted(set(types))


def proportional_targets(counter: Counter[str], total: int, *, minimum: int = 1) -> dict[str, int]:
    keys = list(counter)
    if not keys:
        return {}
    raw = {key: counter[key] * total / sum(counter.values()) for key in keys}
    targets = {key: max(minimum, math.floor(value)) for key, value in raw.items()}
    current = sum(targets.values())
    if current > total:
        ranked = sorted(
            keys,
            key=lambda key: (targets[key] - raw[key], targets[key], key),
            reverse=True,
        )
        for key in ranked:
            while current > total and targets[key] > minimum:
                targets[key] -= 1
                current -= 1
    elif current < total:
        ranked = sorted(keys, key=lambda key: (raw[key] - targets[key], counter[key], key), reverse=True)
        idx = 0
        while current < total:
            key = ranked[idx % len(ranked)]
            targets[key] += 1
            current += 1
            idx += 1
    return targets


def select_candidates(candidates: list[Candidate], sample_size: int = 50) -> list[Candidate]:
    by_id = {candidate.annotation_id: candidate for candidate in candidates}
    selected: list[Candidate] = []
    selected_ids: set[str] = set()

    for annotation_id in CURATED_PRIORITY_IDS:
        candidate = by_id.get(annotation_id)
        if candidate and annotation_id not in selected_ids:
            selected.append(candidate)
            selected_ids.add(annotation_id)

    source_targets = proportional_targets(Counter(candidate.source for candidate in candidates), sample_size)
    stratum_targets = proportional_targets(Counter(candidate.sample_stratum for candidate in candidates), sample_size)
    selected_source_counts = Counter(candidate.source for candidate in selected)
    selected_stratum_counts = Counter(candidate.sample_stratum for candidate in selected)
    selected_type_counts = Counter(tag for candidate in selected for tag in candidate.target_types)
    selected_datasets = {candidate.dataset_name for candidate in selected}

    while len(selected) < sample_size:
        best_candidate: Candidate | None = None
        best_score: tuple[float, float, float, float, str] | None = None
        for candidate in candidates:
            if candidate.annotation_id in selected_ids:
                continue
            tag_score = 0.0
            for tag in candidate.target_types:
                deficit = max(0, TYPE_TARGETS[tag] - selected_type_counts[tag])
                tag_score += deficit * TYPE_WEIGHTS[tag]
            uncovered_score = sum(1.0 for tag in candidate.target_types if selected_type_counts[tag] == 0) * 5.0
            source_score = 2.0 if selected_source_counts[candidate.source] < source_targets[candidate.source] else 0.0
            stratum_score = 1.5 if selected_stratum_counts[candidate.sample_stratum] < stratum_targets[candidate.sample_stratum] else 0.0
            dataset_score = 0.6 if candidate.dataset_name not in selected_datasets else 0.0
            prediction_score = 0.5 if candidate.bert_prob else 0.0
            short_text_bonus = 0.2 if "short_answer" in candidate.target_types else 0.0
            score_tuple = (
                uncovered_score + tag_score + source_score + stratum_score + dataset_score + prediction_score + short_text_bonus,
                source_score + stratum_score,
                prediction_score,
                dataset_score,
                candidate.annotation_id,
            )
            if best_score is None or score_tuple > best_score:
                best_candidate = candidate
                best_score = score_tuple
        if best_candidate is None:
            break
        selected.append(best_candidate)
        selected_ids.add(best_candidate.annotation_id)
        selected_source_counts[best_candidate.source] += 1
        selected_stratum_counts[best_candidate.sample_stratum] += 1
        selected_datasets.add(best_candidate.dataset_name)
        for tag in best_candidate.target_types:
            selected_type_counts[tag] += 1

    selected = sorted(selected, key=lambda item: item.annotation_id)
    if len(selected) != sample_size:
        raise SystemExit(f"Expected {sample_size} samples, got {len(selected)}")
    return selected


def build_column_template_rows(selected: list[Candidate]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for candidate in selected:
        for idx in range(1, TEMPLATE_ROWS_PER_PAIR + 1):
            rows.append(
                {
                    "annotation_id": candidate.annotation_id,
                    "pair_id": candidate.pair_id,
                    "column_id": f"C{idx:02d}",
                    "span_a": "",
                    "span_b": "",
                    "relation_type": "",
                    "relation_strength": "",
                    "alignment_direction": "",
                    "is_core_column": "",
                    "supports_resonance": "",
                    "notes": "",
                }
            )
    return rows


def ensure_output_dir(output_dir: Path, *, overwrite: bool) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    if output_dir.resolve() != artifact_path("formal_300_v1", "diagraph_gold_50").resolve():
        raise SystemExit(
            "This task must write only to experiments/dialogue_syntax_bert/artifacts/formal_300_v1/diagraph_gold_50/"
        )
    if not overwrite:
        return output_dir
    return output_dir


def write_workbook(
    path: Path,
    rows: list[dict[str, str]],
    fieldnames: list[str],
    *,
    overwrite: bool,
    dropdowns: dict[str, list[str]] | None = None,
) -> None:
    output_path = ensure_can_write(path, overwrite=overwrite)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, field in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
        width = 16
        if field in {"turn_a", "turn_b", "annotator_note", "notes"}:
            width = 48
        elif field in {"evidence_span_a", "evidence_span_b", "span_a", "span_b"}:
            width = 26
        elif field in {"source", "dataset_name", "sample_stratum", "relation_type"}:
            width = 20
        sheet.column_dimensions[cell.column_letter].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if dropdowns:
        for field, values in dropdowns.items():
            col_idx = fieldnames.index(field) + 1
            formula = '"' + ",".join(values) + '"'
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.promptTitle = field
            validation.prompt = f"Allowed values: {', '.join(values)}"
            sheet.add_data_validation(validation)
            validation.add(f"{sheet.cell(row=2, column=col_idx).column_letter}2:{sheet.cell(row=max(2, len(rows)+1), column=col_idx).column_letter}{max(2, len(rows)+1)}")

    workbook.save(output_path)


def build_annotation_guide() -> str:
    return """# diagraph_gold_50 标注说明

## 1. 这份标注在评估什么

本轮不是评估“一个 pair 有没有共鸣”本身，而是评估跨句图谱的**纵栏生成质量**。

这里的跨句图谱不是普通知识图谱，也不是实体关系抽取。  
它关注的是：A 句和 B 句之间，哪些成分发生了可解释的映射、复现、替代、回指、对比、填槽或语用功能对应。

## 2. 核心概念

1. 跨句图谱不是普通知识图谱。  
   它不是“实体-属性-关系”的百科式结构。

2. 跨句图谱是话句之间成分映射的结果。  
   每一条纵栏都应对应 A、B 两个话轮中的一组成分或功能单元。

3. pair-level yes/no 只能判断是否有共鸣，不能替代 column-level diagraph。  
   一个 pair 即使已经被判为 `resonance_present=yes`，也仍然需要继续判断：  
   到底有哪些纵栏、它们是什么类型、哪些是核心纵栏。

4. BERT 只能辅助判断 pair 置信度，不能直接生成图谱。  
   `bert_prob` 只是一种 shadow confidence，不是图谱边。

5. 图谱质量最终取决于纵栏映射是否准确。  
   因此本轮标注的重点不是“再判一次 yes/no”，而是把纵栏找准、类型标准、主次分清。

## 3. 什么可以算纵栏

纵栏可以是：

- 词汇复现
- 句法平行
- 语义替代
- 回指 / 指示
- 填槽
- 简短回答
- 对比
- 修正 / repair
- 类比
- 语用功能对应
- 语气词 / 标点 / modal 对应（仅限确有解释价值时）

请注意：

- 不要把**单纯话题相关**标成图谱纵栏。
- “都在谈同一件事”不等于“出现了稳定成分映射”。
- evidence span 必须来自原文，不能自己改写、概括或补写。

## 4. 文件配合方式

1. `diagraph_gold_50_pair_list.*`  
   用来查看 pair 原文、现有 pair-level 标签、evidence span、备注、shadow 置信度。

2. `diagraph_gold_50_column_annotation_template.*`  
   用来逐条填写纵栏标注。  
   同一个 pair 可以对应多行；每一行代表一个候选纵栏。

## 5. column-level 字段说明

### `column_id`

- 建议从 `C01`、`C02`、`C03`... 顺序使用。
- 模板默认每条 pair 预留 5 行，不够时可继续复制新行。

### `span_a` / `span_b`

- 必须直接摘自原文。
- 尽量贴近最小但足够解释的片段。
- 若是回指、代词、指示词，也要保留原词，不要替换成你理解的指代对象。

### `relation_type`

可选值：

- `lexical_reproduction`
- `syntactic_parallelism`
- `semantic_substitution`
- `coreference_or_demonstrative`
- `slot_filling`
- `short_answer`
- `contrast`
- `repair`
- `analogy`
- `pragmatic_function`
- `punctuation_or_modal`
- `other`

如果一个纵栏兼有多种性质，优先填写**最能解释该纵栏成立的主类型**。

### `relation_strength`

- `strong`: 对应关系清楚、证据强、去掉后会明显削弱图谱
- `medium`: 对应较清楚，但依赖上下文判断
- `weak`: 有一定映射迹象，但边界或类型不够稳

### `alignment_direction`

- `A_to_B`: 主要是 A 触发、B 承接或回应
- `B_to_A`: 主要是 B 倒逼回看 A，或 B 明确重构 A 的成分
- `mutual`: 双向映射都很明显

### `is_core_column`

- `1`: 这是该 pair 成立为共鸣的核心纵栏之一
- `0`: 是辅助纵栏，不是主干
- `?`: 一时无法判定

### `supports_resonance`

- `1`: 该纵栏实际支撑 pair-level resonance
- `0`: 有对应痕迹，但不足以支撑共鸣判断
- `?`: 暂时不确定

## 6. 推荐标注流程

1. 先读 `pair_list`，理解 A/B 两句怎么接起来。
2. 只从原文中找可以落到 `span_a` / `span_b` 的映射。
3. 一条纵栏写一行，不要把两个不同关系混在一行里。
4. 先判断是否存在核心纵栏，再补充边缘纵栏。
5. 如果只是话题延续，没有明确映射，不要硬标。

## 7. 常见误判提醒

1. 不能把“同主题”当作“同纵栏”。  
   例如都在谈“高铁”“台湾”，但没有稳定成分对位时，不应强行画栏。

2. 不能把 BERT 高分当作图谱存在的证据。  
   高分只表示该 pair 可能有隐性共鸣，不等于已经找到了 column。

3. 不要把整句整段一次性全贴进去。  
   除非整句确实作为一个整体发生映射，否则优先标最关键的局部 span。

4. 回指和替代要谨慎。  
   若 B 中的“这 / 那 / 他 / 其 / 那个”确实回指 A 中成分，可标；若只是泛指，不应勉强。

## 8. 本轮目标

本轮的目标不是再训练模型，而是沉淀一批可靠的 column-level gold，用来评估未来的自动跨句图谱生成质量。
"""


def build_generation_quality_plan() -> str:
    return f"""# diagraph generation quality plan

## 1. 为什么需要从 pair-level 转向 column-level

pair-level resonance detection 只能回答“这两个话轮之间是否存在共鸣”。  
但跨句图谱生成真正需要回答的是：

1. 共鸣具体落在哪些成分上；
2. 每一条成分映射属于什么关系类型；
3. 哪些纵栏是主干，哪些只是辅助；
4. 整体图谱是否足够稳定、可解释、可复核。

因此，pair-level `yes/no` 只是入口，不是图谱质量终点。

## 2. 当前 hybrid 如何帮助图谱候选筛选

当前 pair-level shadow 系统已经具备三层信息：

1. 规则信号：`rule_any_positive` 及其相关 evidence
2. MacBERT / multi-seed 概率：`bert_prob`
3. hybrid shadow 决策：`{HYBRID_STRATEGY_NAME}`，阈值约 `{HYBRID_THRESHOLD:.6f}`

这套 hybrid 的作用是：

- 保留 rule-positive 的可解释候选；
- 给 rule-negative 但语义上可能共鸣的 pair 提供补召回线索；
- 作为图谱候选筛选和排序辅助，而不是图谱边本身。

也就是说，hybrid 可以帮助我们决定“优先看哪些 pair”，但不能替代纵栏标注。

## 3. 未来如何评估自动图谱生成质量

未来自动 diagraph generation 的评估，建议以人工 column-level gold 为基准，按“候选列、列类型、核心性、整体程度”四层展开。

### 3.1 基础匹配层

- 预测系统输出若干 column：
  - `span_a`
  - `span_b`
  - `relation_type`
  - `relation_strength`
  - `is_core_column`

- 与人工 gold 做对齐时，应至少比较：
  - span 是否匹配
  - A/B 对位是否正确
  - relation_type 是否正确

### 3.2 建议指标

1. `column precision`  
   预测出的纵栏里，有多少能被 gold 接受。

2. `column recall`  
   gold 中应有的纵栏，有多少被系统找回。

3. `relation_type accuracy`  
   在 span 对齐成立的前提下，关系类型是否判断正确。

4. `core_column F1`  
   对核心纵栏的识别能力。  
   这比普通列更重要，因为核心列直接决定图谱主干。

5. `resonance_degree error`  
   用于比较系统输出的整体共鸣强度与人工判断之间的偏差。  
   可先用“核心列数量 + relation_strength 加权”的方式定义一个可比较分值，再计算误差。

## 4. 匹配策略建议

为减少边界争议，评估时建议保留两套口径：

1. 严格口径  
   `span_a` 与 `span_b` 都要求精确或近精确匹配。

2. 宽松口径  
   允许字符级重叠、包含关系或轻微边界偏移。

这样可以区分：

- 真正结构错误
- 仅仅是边界切分偏移

## 5. 暂不让 BERT 直接生成图谱的原因

当前不应让 BERT 直接生成跨句图谱，原因包括：

1. 现有 BERT 任务是 pair-level resonance classification，不是 column generation。
2. 它没有显式的 span grounding supervision。
3. 它没有 relation_type supervision。
4. 它容易把话题相关误判成结构共鸣。
5. 它给出的是概率，不是可审计的映射结构。
6. 当前 gold 主要是 pair-level，column-level gold 还需要先补齐。

因此，现阶段更合理的路线是：

1. 先建设 `diagraph_gold_50` 这样的 column-level 小金标；
2. 用它评估规则候选、模板生成、span 对位质量；
3. 再讨论是否值得做自动图谱生成或半自动建议。

## 6. 建议的下一阶段路线

1. 先完成 `diagraph_gold_50` 人工标注。
2. 统计各 relation_type 的人工分布和一致性。
3. 定义 strict / relaxed 两套 matching 规则。
4. 设计一个 rule-first 的 column candidate generator。
5. 再用本轮 gold 去评估自动候选生成，而不是直接上生成式 BERT。
"""


def validate_selection(selected: list[Candidate]) -> dict[str, Counter[str] | int]:
    if len(selected) != 50:
        raise SystemExit(f"Expected 50 selected samples, got {len(selected)}")
    if any(candidate.resonance_present != "yes" for candidate in selected):
        raise SystemExit("Selection contains non-yes samples.")
    if len({candidate.annotation_id for candidate in selected}) != 50:
        raise SystemExit("Selection contains duplicate annotation_id values.")
    coverage = Counter(tag for candidate in selected for tag in candidate.target_types)
    missing = [tag for tag in TARGET_TYPES if coverage[tag] == 0]
    if missing:
        raise SystemExit(f"Missing target coverage: {missing}")
    return {
        "source_distribution": Counter(candidate.source for candidate in selected),
        "stratum_distribution": Counter(candidate.sample_stratum for candidate in selected),
        "dataset_distribution": Counter(candidate.dataset_name for candidate in selected),
        "type_coverage": coverage,
        "with_bert_prob": sum(1 for candidate in selected if candidate.bert_prob),
        "with_hybrid_pred": sum(1 for candidate in selected if candidate.hybrid_pred),
    }


def print_report(stats: dict[str, Counter[str] | int]) -> None:
    print("diagraph_gold_50 generation complete")
    print("type coverage:")
    type_coverage = stats["type_coverage"]
    assert isinstance(type_coverage, Counter)
    for tag in TARGET_TYPES:
        print(f"  {tag}: {type_coverage[tag]}")
    print("source distribution:")
    source_distribution = stats["source_distribution"]
    assert isinstance(source_distribution, Counter)
    for key, value in source_distribution.most_common():
        print(f"  {key}: {value}")
    print("sample_stratum distribution:")
    stratum_distribution = stats["stratum_distribution"]
    assert isinstance(stratum_distribution, Counter)
    for key, value in stratum_distribution.most_common():
        print(f"  {key}: {value}")
    print(f"bert_prob available: {stats['with_bert_prob']}/50")
    print(f"hybrid_pred available: {stats['with_hybrid_pred']}/50")


def main() -> None:
    args = parse_args()
    base = artifact_path("formal_300_v1")
    output_dir = Path(args.output_dir)
    ensure_output_dir(output_dir, overwrite=args.overwrite)

    candidates = load_sources(base)
    selected = select_candidates(candidates, sample_size=50)
    pair_list_rows = [candidate.to_pair_list_row() for candidate in selected]
    column_template_rows = build_column_template_rows(selected)

    pair_csv_path = output_dir / "diagraph_gold_50_pair_list.csv"
    pair_xlsx_path = output_dir / "diagraph_gold_50_pair_list.xlsx"
    column_csv_path = output_dir / "diagraph_gold_50_column_annotation_template.csv"
    column_xlsx_path = output_dir / "diagraph_gold_50_column_annotation_template.xlsx"
    guide_path = output_dir / "diagraph_gold_50_annotation_guide.md"
    plan_path = output_dir / "diagraph_generation_quality_plan.md"

    write_csv(pair_csv_path, pair_list_rows, PAIR_LIST_FIELDS, overwrite=args.overwrite)
    write_workbook(pair_xlsx_path, pair_list_rows, PAIR_LIST_FIELDS, overwrite=args.overwrite)

    write_csv(column_csv_path, column_template_rows, COLUMN_TEMPLATE_FIELDS, overwrite=args.overwrite)
    write_workbook(
        column_xlsx_path,
        column_template_rows,
        COLUMN_TEMPLATE_FIELDS,
        overwrite=args.overwrite,
        dropdowns={
            "relation_type": RELATION_TYPES,
            "relation_strength": RELATION_STRENGTHS,
            "alignment_direction": ALIGNMENT_DIRECTIONS,
            "is_core_column": TERNARY_VALUES,
            "supports_resonance": TERNARY_VALUES,
        },
    )

    write_text(guide_path, build_annotation_guide(), overwrite=args.overwrite)
    write_text(plan_path, build_generation_quality_plan(), overwrite=args.overwrite)

    stats = validate_selection(selected)
    print_report(stats)


if __name__ == "__main__":
    main()
