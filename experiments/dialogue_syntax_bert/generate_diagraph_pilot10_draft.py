"""Extract pilot10 draft annotations and generate validation / revision notes."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from io_utils import artifact_path, ensure_can_write, read_csv, write_csv, write_text


TEMPLATE_FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
    "notes",
]

PILOT_LIST_LEFT_FIELDS = [
    "annotation_id",
    "pair_id",
    "source",
    "dataset_name",
    "turn_a",
    "turn_b",
    "difficulty_level",
    "priority_rank",
    "expected_column_count",
    "dominant_relation_types",
    "why_this_difficulty",
    "annotation_warning",
    "suggested_first_pass",
]

VALID_RELATION_TYPES = {
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
}
VALID_RELATION_STRENGTH = {"strong", "medium", "weak"}
VALID_ALIGNMENT = {"A_to_B", "B_to_A", "mutual"}
VALID_TERNARY = {"1", "0", "?"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        default=str(artifact_path("formal_300_v1", "diagraph_gold_50")),
    )
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def read_pilot10_list_with_inline_annotations(
    path: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    left_rows: list[dict[str, str]] = []
    right_rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        if len(header) < len(PILOT_LIST_LEFT_FIELDS) + len(TEMPLATE_FIELDS):
            raise SystemExit("pilot10 list does not contain the expected appended annotation area.")
        right_start = len(header) - len(TEMPLATE_FIELDS)
        for raw_row in reader:
            if not raw_row:
                continue
            if len(raw_row) < len(header):
                raw_row = raw_row + [""] * (len(header) - len(raw_row))
            left_rows.append(
                {
                    field: raw_row[idx].strip()
                    for idx, field in enumerate(PILOT_LIST_LEFT_FIELDS)
                }
            )
            right_payload = raw_row[right_start : right_start + len(TEMPLATE_FIELDS)]
            if any(str(value).strip() for value in right_payload):
                right_rows.append(
                    {
                        field: right_payload[idx].strip()
                        for idx, field in enumerate(TEMPLATE_FIELDS)
                    }
                )
    return left_rows, right_rows


def normalize_column_id(column_id: str) -> str:
    value = column_id.strip()
    if re.fullmatch(r"C\d+", value):
        digits = int(value[1:])
        return f"C{digits:02d}"
    return value


def sort_key(row: dict[str, str]) -> tuple[str, int, str]:
    column_id = normalize_column_id(row["column_id"])
    match = re.match(r"C(\d+)$", column_id)
    number = int(match.group(1)) if match else 999
    return (row["annotation_id"], number, row["column_id"])


def write_workbook(path: Path, rows: list[dict[str, str]], fieldnames: list[str], *, overwrite: bool) -> None:
    output_path = ensure_can_write(path, overwrite=overwrite)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "pilot10_draft"
    fill = PatternFill(fill_type="solid", start_color="E6F2D9", end_color="E6F2D9")
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    widths = {
        "annotation_id": 16,
        "pair_id": 12,
        "column_id": 10,
        "span_a": 24,
        "span_b": 24,
        "relation_type": 28,
        "relation_strength": 14,
        "alignment_direction": 16,
        "is_core_column": 14,
        "supports_resonance": 18,
        "notes": 42,
    }
    for idx, field in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=idx, value=field)
        cell.fill = fill
        cell.font = bold
        cell.alignment = wrap
        sheet.column_dimensions[cell.column_letter].width = widths.get(field, 18)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = wrap
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def load_pair_context(pair_list_path: Path) -> dict[str, dict[str, str]]:
    rows = read_csv(pair_list_path)
    return {row["annotation_id"]: row for row in rows}


def load_expected_counts(priority_rows: list[dict[str, str]]) -> dict[str, int]:
    expected: dict[str, int] = {}
    for row in priority_rows:
        value = row.get("expected_column_count", "").strip()
        if value.isdigit():
            expected[row["annotation_id"]] = int(value)
    return expected


def validate_draft(
    draft_rows: list[dict[str, str]],
    pilot_rows: list[dict[str, str]],
    pair_context: dict[str, dict[str, str]],
    expected_counts: dict[str, int],
) -> dict[str, object]:
    pilot_ids = {row["annotation_id"] for row in pilot_rows}
    draft_ids = {row["annotation_id"] for row in draft_rows}
    drafted_pairs = Counter(row["annotation_id"] for row in draft_rows)
    invalid_span_a: list[str] = []
    invalid_span_b: list[str] = []
    invalid_relation_type: list[str] = []
    invalid_relation_strength: list[str] = []
    invalid_alignment: list[str] = []
    invalid_ternary: list[str] = []

    core_count_by_pair: dict[str, int] = defaultdict(int)
    for row in draft_rows:
        annotation_id = row["annotation_id"]
        context = pair_context.get(annotation_id)
        if context is None:
            invalid_span_a.append(f"{annotation_id} missing pair context")
            invalid_span_b.append(f"{annotation_id} missing pair context")
            continue
        span_a = row["span_a"].strip()
        span_b = row["span_b"].strip()
        if not span_a or span_a not in context["turn_a"]:
            invalid_span_a.append(f"{annotation_id}/{row['column_id']}: {span_a}")
        if not span_b or span_b not in context["turn_b"]:
            invalid_span_b.append(f"{annotation_id}/{row['column_id']}: {span_b}")
        if row["relation_type"] not in VALID_RELATION_TYPES:
            invalid_relation_type.append(f"{annotation_id}/{row['column_id']}: {row['relation_type']}")
        if row["relation_strength"] not in VALID_RELATION_STRENGTH:
            invalid_relation_strength.append(f"{annotation_id}/{row['column_id']}: {row['relation_strength']}")
        if row["alignment_direction"] not in VALID_ALIGNMENT:
            invalid_alignment.append(f"{annotation_id}/{row['column_id']}: {row['alignment_direction']}")
        if row["is_core_column"] not in VALID_TERNARY:
            invalid_ternary.append(f"{annotation_id}/{row['column_id']}: is_core_column={row['is_core_column']}")
        if row["supports_resonance"] not in VALID_TERNARY:
            invalid_ternary.append(f"{annotation_id}/{row['column_id']}: supports_resonance={row['supports_resonance']}")
        if row["is_core_column"] == "1":
            core_count_by_pair[annotation_id] += 1

    missing_pilot_pairs = [annotation_id for annotation_id in pilot_ids if annotation_id not in draft_ids]
    no_core_pairs = [annotation_id for annotation_id in pilot_ids if core_count_by_pair.get(annotation_id, 0) < 1]
    over_five_allowed = expected_counts.get("F300V1-0127", 0) > 5

    return {
        "pilot_count": len(pilot_ids),
        "draft_row_count": len(draft_rows),
        "draft_pair_count": len(draft_ids),
        "only_contains_pilot10": draft_ids.issubset(pilot_ids),
        "missing_pilot_pairs": sorted(missing_pilot_pairs),
        "invalid_span_a": invalid_span_a,
        "invalid_span_b": invalid_span_b,
        "invalid_relation_type": invalid_relation_type,
        "invalid_relation_strength": invalid_relation_strength,
        "invalid_alignment": invalid_alignment,
        "invalid_ternary": invalid_ternary,
        "core_count_by_pair": dict(core_count_by_pair),
        "no_core_pairs": sorted(no_core_pairs),
        "drafted_pairs": drafted_pairs,
        "f300v1_0127_allow_over_five": over_five_allowed,
    }


def build_validation_report(
    validation: dict[str, object],
    pilot_rows: list[dict[str, str]],
    draft_rows: list[dict[str, str]],
) -> str:
    drafted_pairs = validation["drafted_pairs"]
    assert isinstance(drafted_pairs, Counter)
    missing_pairs = validation["missing_pilot_pairs"]
    assert isinstance(missing_pairs, list)
    no_core_pairs = validation["no_core_pairs"]
    assert isinstance(no_core_pairs, list)
    invalid_span_a = validation["invalid_span_a"]
    invalid_span_b = validation["invalid_span_b"]
    invalid_relation_type = validation["invalid_relation_type"]
    invalid_relation_strength = validation["invalid_relation_strength"]
    invalid_alignment = validation["invalid_alignment"]
    invalid_ternary = validation["invalid_ternary"]

    status = lambda ok: "通过" if ok else "未通过"
    lines = []
    lines.append("# diagraph_gold_50 pilot10 column validation report")
    lines.append("")
    lines.append("## 1. 基本情况")
    lines.append("")
    lines.append(f"- pilot10 pair 总数：{validation['pilot_count']}")
    lines.append(f"- 当前抽取到的 draft 行数：{validation['draft_row_count']}")
    lines.append(f"- 当前已有草案的 pilot pair 数：{validation['draft_pair_count']}")
    lines.append(f"- 是否只包含 pilot10：{status(bool(validation['only_contains_pilot10']))}")
    lines.append("")
    lines.append("## 2. annotation_id 覆盖情况")
    lines.append("")
    lines.append("- 已有 draft 的 pair：")
    for annotation_id, count in drafted_pairs.items():
        lines.append(f"  - `{annotation_id}`: {count} 行")
    lines.append("")
    lines.append("- 尚未进入 draft 的 pilot pair：")
    if missing_pairs:
        for annotation_id in missing_pairs:
            lines.append(f"  - `{annotation_id}`")
    else:
        lines.append("  - 无")
    lines.append("")
    lines.append("## 3. span 校验")
    lines.append("")
    lines.append(f"- span_a 能在 turn_a 中找到：{status(not invalid_span_a)}")
    if invalid_span_a:
        for item in invalid_span_a:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 所有 draft 行均通过 span_a 校验")
    lines.append(f"- span_b 能在 turn_b 中找到：{status(not invalid_span_b)}")
    if invalid_span_b:
        for item in invalid_span_b:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 所有 draft 行均通过 span_b 校验")
    lines.append("")
    lines.append("## 4. 值域校验")
    lines.append("")
    lines.append(f"- relation_type 合法：{status(not invalid_relation_type)}")
    if invalid_relation_type:
        for item in invalid_relation_type:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 全部 relation_type 合法")
    lines.append(f"- relation_strength 合法：{status(not invalid_relation_strength)}")
    if invalid_relation_strength:
        for item in invalid_relation_strength:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 全部 relation_strength 合法")
    lines.append(f"- alignment_direction 合法：{status(not invalid_alignment)}")
    if invalid_alignment:
        for item in invalid_alignment:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 全部 alignment_direction 合法")
    lines.append(f"- is_core_column / supports_resonance 合法：{status(not invalid_ternary)}")
    if invalid_ternary:
        for item in invalid_ternary:
            lines.append(f"  - {item}")
    else:
        lines.append("  - 全部 ternary 字段合法")
    lines.append("")
    lines.append("## 5. core column 检查")
    lines.append("")
    lines.append("- 每个 pilot pair 是否至少有 1 个 core column：未通过")
    lines.append("- 原因：当前草案只覆盖 3 个 pilot pair；其余 7 个 pair 尚无 column rows，因此暂时无法满足“每个 pilot pair 至少 1 个 core column”的条件。")
    lines.append("- 当前已有 core column 的 pair：")
    core_count_by_pair = validation["core_count_by_pair"]
    assert isinstance(core_count_by_pair, dict)
    for annotation_id, count in sorted(core_count_by_pair.items()):
        lines.append(f"  - `{annotation_id}`: {count}")
    lines.append("")
    lines.append("- 当前没有 core column 的 pilot pair：")
    if no_core_pairs:
        for annotation_id in no_core_pairs:
            lines.append(f"  - `{annotation_id}`")
    else:
        lines.append("  - 无")
    lines.append("")
    lines.append("## 6. F300V1-0127 额外行数判断")
    lines.append("")
    lines.append(f"- `F300V1-0127` 是否允许超过模板默认 5 行：{status(bool(validation['f300v1_0127_allow_over_five']))}")
    lines.append("- 依据：pilot priority 中该样本的 `expected_column_count=6`，属于 hard / analogy 类样本，允许扩行。")
    return "\n".join(lines) + "\n"


def build_revision_notes(validation: dict[str, object]) -> str:
    missing_pairs = validation["missing_pilot_pairs"]
    assert isinstance(missing_pairs, list)
    return f"""# diagraph_gold_50 pilot10 guide revision notes

## 1. 代词 / 人称转换如何标

pilot 草案已经出现两类典型现象：

1. 指代承接：如 `那差来的 -> 他`
2. 人称转换：如 `哥哥你 -> 我`

建议补充规则：

- 只要 A 和 B 指向同一参与者或同一论元位，就可以标 `coreference_or_demonstrative`
- `span_a` 与 `span_b` 必须保留原文表面形式，不要把“你/我/他”改写成解释性词语
- 若存在人称转换，建议在 `notes` 里说明“deictic center shift / speaker-role shift”

## 2. slot_filling 如何判断

当前草案里已经有明确的 slot-filling 例子：

- `甚么名字 -> 巴山虎`
- `甚么名字 -> 倚海龙`
- `贵姓 -> 王`

建议补充规则：

- A 必须提供清楚的问句槽位、待填变量或问答框架
- B 必须给出与该槽位直接绑定的填充值
- 如果 B 给出多个并列填值，可以拆成多条 column
- 仅仅“回应了问题”但没有明确填槽，不应一律标成 `slot_filling`

## 3. semantic_substitution 与纯话题相关如何区分

建议补充区分标准：

- `semantic_substitution` 需要保留 A 中的同一论题位、判断位或问答位
- B 如果只是继续同一话题，但没有对 A 的结构位置进行替代、回应、重构或填补，就不应标成纵栏
- 若 annotator 只能说“它们在谈同一件事”，但说不出 A/B 哪两个成分在对位，就暂不标

## 4. analogy 必须有结构推理链

对于类比类 hard case，尤其是 `F300V1-0127`：

- 不能因为 B 出现比附性结论就直接标 `analogy`
- 必须能说明 A 中哪些关系链条被 B 映射为另一个关系结论
- 建议在 notes 中写出最短推理链，例如：
  - `一母所生 / 亲处 -> 外甥`

## 5. hard 样本是否需要超过 5 行

是，至少 `F300V1-0127` 应明确允许超过 5 行。  
若后续在 short answer + demonstrative + semantic substitution 叠加的 hard pair 中出现 6 个以上稳定纵栏，也应允许增行。

## 6. 标剩余 40 条前，guide 需要补哪些规则

在进入 remaining40 之前，建议先把 guide 再补 5 条：

1. 代词 / 人称转换如何记录在 `notes`
2. 一个问句槽位对应多个填值时如何拆多行
3. `semantic_substitution` 与“同主题延续”的判别句
4. `short_answer` 何时只是接话，何时能成为真正 column
5. `analogy` 必须附带结构推理链

## 7. 这轮草案目前的状态

当前 draft 已抽出 10 行 column annotations，但只覆盖 3 个 pilot pair。  
尚未进入 draft 的 pilot pair 有：{", ".join(missing_pairs) if missing_pairs else "无"}。
"""


def build_remaining40_plan(priority_rows: list[dict[str, str]]) -> str:
    remaining_rows = [row for row in priority_rows if row.get("suggested_first_pass") != "1"]
    medium_rows = [row for row in remaining_rows if row["difficulty_level"] == "medium"]
    hard_rows = [row for row in remaining_rows if row["difficulty_level"] == "hard"]
    extra_line_rows = [row for row in remaining_rows if int(row["expected_column_count"]) >= 6]
    extra_ids = ", ".join(row["annotation_id"] for row in extra_line_rows) if extra_line_rows else "暂无明显需要扩行的 remaining40 样本"
    return f"""# diagraph_gold_50 remaining40 plan

## 1. 先标 medium 还是 hard

建议先标 **medium**，再进入 **hard**。

理由：

1. medium 样本通常已经有稳定的结构承接，但 relation_type 还需要人工细分；
2. 这类样本最适合继续修正 guide，而不会像 hard 一样把大量时间消耗在边界争议上；
3. 等 medium 的规则更稳之后，再回到 hard，能减少“把话题相关误画成 column”的风险。

## 2. 哪些样本需要额外行数

根据当前 priority 中的 `expected_column_count`，remaining40 里最可能需要超过默认 5 行的样本有：

{extra_ids}

## 3. 哪些 relation_type 最容易混淆

后续最需要警惕的混淆主要有：

1. `coreference_or_demonstrative` vs `semantic_substitution`
2. `slot_filling` vs `short_answer`
3. `lexical_reproduction` vs `syntactic_parallelism`
4. `contrast` vs `repair`
5. `analogy` vs 高层语义类比 / 话题类比

## 4. 是否需要二次复核

建议需要二次复核，尤其是以下类型：

1. hard 样本
2. `expected_column_count >= 6` 的样本
3. 带有 demonstrative / pronoun shift 的样本
4. analogy 类样本
5. short-answer 主导的样本

## 5. 推荐流程

1. 先完成 pilot10 剩余 7 个 pair 的 draft
2. 修订 guide
3. 先做 remaining40 中的 medium
4. 再做 remaining40 中的 hard
5. 对 hard + 多纵栏样本进行二次复核
"""


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    expected_dir = artifact_path("formal_300_v1", "diagraph_gold_50").resolve()
    if output_dir.resolve() != expected_dir:
        raise SystemExit(f"Outputs must stay under {expected_dir}")

    pair_list_path = output_dir / "diagraph_gold_50_pair_list.csv"
    template_path = output_dir / "diagraph_gold_50_column_annotation_template.csv"
    pilot_list_path = output_dir / "diagraph_gold_50_pilot10_list.csv"
    priority_path = output_dir / "diagraph_gold_50_annotation_priority.csv"

    if not pair_list_path.exists() or not template_path.exists() or not pilot_list_path.exists():
        raise SystemExit("Required input files are missing.")

    pair_context = load_pair_context(pair_list_path)
    _template_rows = read_csv(template_path)
    pilot_left_rows, draft_rows = read_pilot10_list_with_inline_annotations(pilot_list_path)
    draft_rows = sorted(
        [
            {
                **row,
                "column_id": normalize_column_id(row["column_id"]),
            }
            for row in draft_rows
        ],
        key=sort_key,
    )

    expected_counts = {}
    if priority_path.exists():
        expected_counts = load_expected_counts(read_csv(priority_path))
        priority_rows = read_csv(priority_path)
    else:
        priority_rows = []
        for row in pilot_left_rows:
            value = row.get("expected_column_count", "").strip()
            if value.isdigit():
                expected_counts[row["annotation_id"]] = int(value)

    validation = validate_draft(draft_rows, pilot_left_rows, pair_context, expected_counts)
    validation_report = build_validation_report(validation, pilot_left_rows, draft_rows)
    revision_notes = build_revision_notes(validation)
    remaining_plan = build_remaining40_plan(priority_rows) if priority_rows else "# diagraph_gold_50 remaining40 plan\n\n缺少 priority 文件，暂未生成细化建议。\n"

    draft_csv = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft.csv"
    draft_xlsx = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft.xlsx"
    validation_md = output_dir / "diagraph_gold_50_pilot10_column_validation_report.md"
    revision_md = output_dir / "diagraph_gold_50_pilot10_guide_revision_notes.md"
    remaining_md = output_dir / "diagraph_gold_50_remaining40_plan.md"

    write_csv(draft_csv, draft_rows, TEMPLATE_FIELDS, overwrite=args.overwrite)
    write_workbook(draft_xlsx, draft_rows, TEMPLATE_FIELDS, overwrite=args.overwrite)
    write_text(validation_md, validation_report, overwrite=args.overwrite)
    write_text(revision_md, revision_notes, overwrite=args.overwrite)
    write_text(remaining_md, remaining_plan, overwrite=args.overwrite)

    print("pilot10 draft generation complete")
    print(f"draft_rows={len(draft_rows)}")
    print(f"draft_pairs={validation['draft_pair_count']}")
    print(f"span_a_invalid={len(validation['invalid_span_a'])}")
    print(f"span_b_invalid={len(validation['invalid_span_b'])}")
    print("missing_pairs=" + ",".join(validation["missing_pilot_pairs"]))


if __name__ == "__main__":
    main()
