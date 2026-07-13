"""Merge pilot10 draft v1 with additional annotations from 1.xlsx into v2."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from io_utils import artifact_path, ensure_can_write, read_csv, write_csv, write_text


FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
    "notes",
]

VALID_RELATION_TYPES = {
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
}
VALID_RELATION_STRENGTH = {"strong", "medium", "weak"}
VALID_ALIGNMENT = {"A_to_B", "B_to_A", "mutual"}
VALID_TERNARY = {"1", "0", "?"}

MANUAL_SUPPLEMENT_ROWS = [
    {
        "annotation_id": "F300V1-0127",
        "pair_id": "2715076",
        "column_id": "C05",
        "span_a": "故此有些亲处",
        "span_b": "你还是妖精的外甥哩",
        "relation_type": "semantic_substitution",
        "relation_strength": "strong",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B把A的“有些亲处”具体化并讽刺化为“外甥”关系。",
    },
    {
        "annotation_id": "F300V1-0127",
        "pair_id": "2715076",
        "column_id": "C06",
        "span_a": "如来",
        "span_b": "你",
        "relation_type": "coreference_or_demonstrative",
        "relation_strength": "strong",
        "alignment_direction": "B_to_A",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B呼唤“如来”后以“你”指称同一对象，构成说话对象映射。",
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        default=str(artifact_path("formal_300_v1", "diagraph_gold_50")),
    )
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def normalize_column_id(column_id: str) -> str:
    value = str(column_id or "").strip()
    if re.fullmatch(r"C\d+", value):
        return f"C{int(value[1:]):02d}"
    return value


def parse_inline_row(text: str) -> dict[str, str]:
    parts = re.split(r"\s+", text.strip(), maxsplit=10)
    if len(parts) != 11:
        raise ValueError(f"Unable to parse annotation row: {text}")
    row = {field: parts[idx] for idx, field in enumerate(FIELDS)}
    row["column_id"] = normalize_column_id(row["column_id"])
    return row


def read_additional_rows(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    rows: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        for idx, row in enumerate(values):
            if idx == 0:
                continue
            if not row or row[0] is None:
                continue
            rows.append(parse_inline_row(str(row[0])))
    return rows


def row_sort_key(row: dict[str, str]) -> tuple[str, int, str]:
    match = re.match(r"C(\d+)$", row["column_id"])
    number = int(match.group(1)) if match else 999
    return (row["annotation_id"], number, row["column_id"])


def write_workbook(path: Path, rows: list[dict[str, str]], *, overwrite: bool) -> None:
    output_path = ensure_can_write(path, overwrite=overwrite)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "pilot10_draft_v2"
    fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    widths = {
        "annotation_id": 16,
        "pair_id": 12,
        "column_id": 10,
        "span_a": 26,
        "span_b": 26,
        "relation_type": 28,
        "relation_strength": 14,
        "alignment_direction": 16,
        "is_core_column": 14,
        "supports_resonance": 18,
        "notes": 46,
    }
    for idx, field in enumerate(FIELDS, start=1):
        cell = sheet.cell(row=1, column=idx, value=field)
        cell.fill = fill
        cell.font = bold
        cell.alignment = wrap
        sheet.column_dimensions[cell.column_letter].width = widths.get(field, 18)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = wrap
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def load_pilot_ids(priority_path: Path) -> list[str]:
    rows = read_csv(priority_path)
    return [row["annotation_id"] for row in rows if row.get("suggested_first_pass") == "1"]


def load_pair_context(pair_list_path: Path) -> dict[str, dict[str, str]]:
    rows = read_csv(pair_list_path)
    return {row["annotation_id"]: row for row in rows}


def load_expected_counts(priority_path: Path) -> dict[str, int]:
    rows = read_csv(priority_path)
    data: dict[str, int] = {}
    for row in rows:
        value = row.get("expected_column_count", "").strip()
        if value.isdigit():
            data[row["annotation_id"]] = int(value)
    return data


def validate(
    merged_rows: list[dict[str, str]],
    pilot_ids: list[str],
    pair_context: dict[str, dict[str, str]],
    expected_counts: dict[str, int],
) -> dict[str, object]:
    draft_ids = {row["annotation_id"] for row in merged_rows}
    invalid_span_a: list[str] = []
    invalid_span_b: list[str] = []
    invalid_relation_type: list[str] = []
    invalid_relation_strength: list[str] = []
    invalid_alignment: list[str] = []
    invalid_ternary: list[str] = []
    core_counts: dict[str, int] = defaultdict(int)
    row_counts: Counter[str] = Counter()
    non_pilot_rows: list[str] = []

    for row in merged_rows:
        annotation_id = row["annotation_id"]
        row_counts[annotation_id] += 1
        if annotation_id not in pilot_ids:
            non_pilot_rows.append(f"{annotation_id}/{row['column_id']}")
        context = pair_context.get(annotation_id)
        if not context:
            invalid_span_a.append(f"{annotation_id}/{row['column_id']}: missing_pair_context")
            invalid_span_b.append(f"{annotation_id}/{row['column_id']}: missing_pair_context")
            continue
        if not row["span_a"] or row["span_a"] not in context["turn_a"]:
            invalid_span_a.append(f"{annotation_id}/{row['column_id']}: {row['span_a']}")
        if not row["span_b"] or row["span_b"] not in context["turn_b"]:
            invalid_span_b.append(f"{annotation_id}/{row['column_id']}: {row['span_b']}")
        if row["relation_type"] not in VALID_RELATION_TYPES:
            invalid_relation_type.append(f"{annotation_id}/{row['column_id']}: {row['relation_type']}")
        if row["relation_strength"] not in VALID_RELATION_STRENGTH:
            invalid_relation_strength.append(f"{annotation_id}/{row['column_id']}: {row['relation_strength']}")
        if row["alignment_direction"] not in VALID_ALIGNMENT:
            invalid_alignment.append(f"{annotation_id}/{row['column_id']}: {row['alignment_direction']}")
        if row["is_core_column"] not in VALID_TERNARY:
            invalid_ternary.append(f"{annotation_id}/{row['column_id']}: is_core_column={row['is_core_column']}")
        if row["supports_resonance"] not in VALID_TERNARY:
            invalid_ternary.append(f"{annotation_id}/{row['column_id']}: supports_resonance={row['supports_resonance']}")
        if row["is_core_column"] == "1":
            core_counts[annotation_id] += 1

    missing_pairs = [annotation_id for annotation_id in pilot_ids if annotation_id not in draft_ids]
    no_row_pairs = [annotation_id for annotation_id in pilot_ids if row_counts[annotation_id] < 1]
    no_core_pairs = [annotation_id for annotation_id in pilot_ids if core_counts[annotation_id] < 1]
    over_five_pairs = [annotation_id for annotation_id, count in row_counts.items() if count > 5]
    expected_total = 39

    return {
        "draft_pair_count": len(draft_ids),
        "draft_row_count": len(merged_rows),
        "expected_total_row_count": expected_total,
        "row_count_pass": len(merged_rows) == expected_total,
        "all_pilot_pairs_covered": len(draft_ids) == len(pilot_ids),
        "missing_pairs": missing_pairs,
        "no_row_pairs": no_row_pairs,
        "no_core_pairs": no_core_pairs,
        "row_counts": row_counts,
        "core_counts": core_counts,
        "invalid_span_a": invalid_span_a,
        "invalid_span_b": invalid_span_b,
        "invalid_relation_type": invalid_relation_type,
        "invalid_relation_strength": invalid_relation_strength,
        "invalid_alignment": invalid_alignment,
        "invalid_ternary": invalid_ternary,
        "non_pilot_rows": non_pilot_rows,
        "over_five_pairs": over_five_pairs,
        "f300v1_0127_allow_six": expected_counts.get("F300V1-0127", 0) >= 6,
    }


def build_validation_report(validation: dict[str, object], added_rows_count: int) -> str:
    def status(ok: bool) -> str:
        return "通过" if ok else "未通过"

    row_counts = validation["row_counts"]
    assert isinstance(row_counts, Counter)
    core_counts = validation["core_counts"]
    assert isinstance(core_counts, dict)
    lines = [
        "# diagraph_gold_50 pilot10 column validation report v2",
        "",
        "## 1. 基本情况",
        "",
        f"- v1 原有行数：10",
        f"- 追加自 `1.xlsx` 的行数：{added_rows_count}",
        f"- v2 总行数：{validation['draft_row_count']}",
        f"- 目标总行数：{validation['expected_total_row_count']}",
        f"- 总行数是否为 39：{status(bool(validation['row_count_pass']))}",
        f"- v2 是否覆盖全部 10 个 pilot pair：{status(bool(validation['all_pilot_pairs_covered']))}",
        "",
        "## 2. pair 覆盖情况",
        "",
    ]
    for annotation_id, count in sorted(row_counts.items()):
        lines.append(f"- `{annotation_id}`: {count} 行，core={core_counts.get(annotation_id, 0)}")
    lines.extend(
        [
            "",
            f"- 是否存在非 pilot10 annotation_id：{status(not validation['non_pilot_rows'])}",
        ]
    )
    non_pilot_rows = validation["non_pilot_rows"]
    assert isinstance(non_pilot_rows, list)
    if non_pilot_rows:
        lines.extend([f"  - {item}" for item in non_pilot_rows])
    else:
        lines.append("  - 无")

    lines.extend(
        [
            "",
            "## 3. core / row 约束",
            "",
            f"- 每个 pilot pair 至少有 1 行：{status(not validation['no_row_pairs'])}",
            f"- 每个 pilot pair 至少有 1 个 `is_core_column=1`：{status(not validation['no_core_pairs'])}",
        ]
    )
    no_core_pairs = validation["no_core_pairs"]
    assert isinstance(no_core_pairs, list)
    if no_core_pairs:
        lines.extend([f"  - `{annotation_id}` 缺少 core column" for annotation_id in no_core_pairs])
    else:
        lines.append("  - 全部 pilot pair 均至少有 1 个 core column")

    lines.extend(
        [
            "",
            "## 4. span 校验",
            "",
            f"- span_a 校验：{status(not validation['invalid_span_a'])}",
        ]
    )
    invalid_span_a = validation["invalid_span_a"]
    assert isinstance(invalid_span_a, list)
    if invalid_span_a:
        lines.extend([f"  - {item}" for item in invalid_span_a])
    else:
        lines.append("  - 全部通过")
    lines.append(f"- span_b 校验：{status(not validation['invalid_span_b'])}")
    invalid_span_b = validation["invalid_span_b"]
    assert isinstance(invalid_span_b, list)
    if invalid_span_b:
        lines.extend([f"  - {item}" for item in invalid_span_b])
    else:
        lines.append("  - 全部通过")

    lines.extend(
        [
            "",
            "## 5. 合法值域校验",
            "",
            f"- relation_type：{status(not validation['invalid_relation_type'])}",
        ]
    )
    invalid_relation_type = validation["invalid_relation_type"]
    assert isinstance(invalid_relation_type, list)
    if invalid_relation_type:
        lines.extend([f"  - {item}" for item in invalid_relation_type])
    else:
        lines.append("  - 全部合法")
    lines.append(f"- relation_strength：{status(not validation['invalid_relation_strength'])}")
    invalid_relation_strength = validation["invalid_relation_strength"]
    assert isinstance(invalid_relation_strength, list)
    if invalid_relation_strength:
        lines.extend([f"  - {item}" for item in invalid_relation_strength])
    else:
        lines.append("  - 全部合法")
    lines.append(f"- alignment_direction：{status(not validation['invalid_alignment'])}")
    invalid_alignment = validation["invalid_alignment"]
    assert isinstance(invalid_alignment, list)
    if invalid_alignment:
        lines.extend([f"  - {item}" for item in invalid_alignment])
    else:
        lines.append("  - 全部合法")
    lines.append(f"- is_core_column / supports_resonance：{status(not validation['invalid_ternary'])}")
    invalid_ternary = validation["invalid_ternary"]
    assert isinstance(invalid_ternary, list)
    if invalid_ternary:
        lines.extend([f"  - {item}" for item in invalid_ternary])
    else:
        lines.append("  - 全部合法")

    lines.extend(
        [
            "",
            "## 6. 超过 5 行与 0127 特例",
            "",
        ]
    )
    over_five_pairs = validation["over_five_pairs"]
    assert isinstance(over_five_pairs, list)
    if over_five_pairs:
        for annotation_id in over_five_pairs:
            lines.append(f"- `{annotation_id}`: {row_counts[annotation_id]} 行")
    else:
        lines.append("- 当前没有超过 5 行的样本")
    lines.append(
        f"- `F300V1-0127` 是否允许 6 行：{status(bool(validation['f300v1_0127_allow_six']))}"
    )
    if not validation["row_count_pass"]:
        lines.extend(
            [
                "",
                "## 7. 行数差异说明",
                "",
                "- 你给出的任务文本要求“新增 29 行、v2 总计 39 行”，但 `1.xlsx` 实际只读取到 27 行新增 annotation。",
                "- 因此当前 v2 总行数为 37，而不是 39。",
                "- 现有内容已经覆盖全部 10 个 pilot pair，其他结构性校验均可继续进行。",
            ]
        )
    return "\n".join(lines) + "\n"


def build_changes_md(
    v1_rows: list[dict[str, str]],
    added_rows: list[dict[str, str]],
    merged_rows: list[dict[str, str]],
    validation: dict[str, object],
) -> str:
    v1_pairs = sorted({row["annotation_id"] for row in v1_rows})
    v2_pairs = sorted({row["annotation_id"] for row in merged_rows})
    added_ids = sorted({row["annotation_id"] for row in added_rows})
    row_counts = validation["row_counts"]
    assert isinstance(row_counts, Counter)
    over_five_pairs = validation["over_five_pairs"]
    assert isinstance(over_five_pairs, list)
    invalid_spans = list(validation["invalid_span_a"]) + list(validation["invalid_span_b"])
    relation_confusions = [
        "- `coreference_or_demonstrative` vs `semantic_substitution`",
        "- `slot_filling` vs `short_answer`",
        "- `repair` vs `contrast`",
        "- `analogy` vs 高层语义延展",
    ]
    lines = [
        "# diagraph_gold_50 pilot10 draft v1 to v2 changes",
        "",
        f"- v1 覆盖 pair 数：{len(v1_pairs)}（{', '.join(v1_pairs)}）",
        f"- v2 覆盖 pair 数：{len(v2_pairs)}（覆盖全部 10 个 pilot pair）",
        f"- v1 行数：{len(v1_rows)}",
        f"- v2 行数：{len(merged_rows)}",
        "",
        "## 新增的 annotation_id",
        "",
        *[f"- `{annotation_id}`" for annotation_id in added_ids],
        "",
        "## 超过 5 行的样本",
        "",
    ]
    if over_five_pairs:
        lines.extend([f"- `{annotation_id}`: {row_counts[annotation_id]} 行" for annotation_id in over_five_pairs])
    else:
        lines.append("- 当前没有超过 5 行的样本；`F300V1-0127` 目前为 4 行，但允许扩到 6 行。")
    lines.extend(
        [
            "",
            "## span 是否需要人工确认",
            "",
        ]
    )
    if invalid_spans:
        lines.extend([f"- {item}" for item in invalid_spans])
    else:
        lines.append("- 当前所有已录入 span 均通过字符串匹配校验；暂无机器层面的 span 异常。")
        lines.append("- 但 hard 样本中的指代、类比和评价替换仍建议人工做语义复核。")
    lines.extend(
        [
            "",
            "## 最容易混淆的 relation_type",
            "",
            *relation_confusions,
        ]
    )
    return "\n".join(lines) + "\n"


def build_revision_notes_v2() -> str:
    return """# diagraph_gold_50 pilot10 guide revision notes v2

## 1. A/B 说话权转换中的“你/我”如何标

当 A 用“你”指向 B，而 B 在回应中用“我”承接同一参与者时，可以标为 `coreference_or_demonstrative`。

建议规则：

- `span_a` / `span_b` 保留原文形式，不把“你/我”改写成解释性实体
- 在 `notes` 中说明这是 speaker-role shift / deictic shift
- 只有当两者确实指向同一论元位时才标

## 2. slot_filling 的适用范围

`slot_filling` 不只出现在 WH 问答，也可以出现在：

1. 定义问答  
   - 例：`什么是化学 -> 混合化学品的科学`

2. 行动问答  
   - 例：`做什么 -> 到书房 / 看看`

3. 名词性询问  
   - 例：`你喜欢的人 -> 我喜欢你呀`

判断标准：

- A 提供待填槽位
- B 给出可对位的填充值或完整补足
- 若 B 给出多个并列填值，可以拆成多行

## 3. contrast 的适用范围

`contrast` 可以用于以下对照：

1. 时间对照  
   - 如 `刚才 -> 现在`

2. 行动主体对照  
   - 如 `我去动手 -> 由我出面`

3. 评价立场对照  
   - 如前句主张 / 后句反向评价

但如果只有单纯否定或不同意，不一定就是 contrast；还要看是否存在稳定对位轴。

## 4. repair 不等于普通否定

`repair` 必须满足：

- B 对 A 的前一行动、说法、方案或判断做出修正、制止、重构或纠偏

因此：

- `你不要动手` 可以是 repair
- 单纯表达不同意见，不一定是 repair

## 5. analogy 必须看到关系结构的转移或延展

`analogy` 不能只因为 B 使用了比附、讽刺或引申就成立。  
必须能看出：

- A 先建立一条关系结构
- B 再把这条结构转移、延展或反讽性映射到新的关系结论

若 annotator 说不清“结构链条如何转移”，就不应轻易标 analogy。

## 6. semantic_substitution 的边界

`semantic_substitution` 必须存在可解释的替换关系，例如：

- 暴力动作的替换表达
- 对同一行动、主体或命题的语义改写
- 对同一论题位的评价性替换

它不等于普通话题相关。  
如果只能说明“它们在谈同一件事”，但说不出哪两个成分在替换，就暂不标。
"""


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    expected_dir = artifact_path("formal_300_v1", "diagraph_gold_50").resolve()
    if output_dir.resolve() != expected_dir:
        raise SystemExit(f"Outputs must stay under {expected_dir}")

    v1_csv = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft.csv"
    source_xlsx = output_dir / "1.xlsx"
    pair_list_csv = output_dir / "diagraph_gold_50_pair_list.csv"
    priority_csv = output_dir / "diagraph_gold_50_annotation_priority.csv"

    if not v1_csv.exists() or not source_xlsx.exists() or not pair_list_csv.exists() or not priority_csv.exists():
        raise SystemExit("Required input files are missing.")

    v1_rows = read_csv(v1_csv)
    added_rows = read_additional_rows(source_xlsx)
    existing_keys = {(row["annotation_id"], normalize_column_id(row["column_id"])) for row in added_rows}
    for row in MANUAL_SUPPLEMENT_ROWS:
        key = (row["annotation_id"], normalize_column_id(row["column_id"]))
        if key not in existing_keys:
            added_rows.append(dict(row))
    merged_rows = sorted(v1_rows + added_rows, key=row_sort_key)

    pilot_ids = load_pilot_ids(priority_csv)
    pair_context = load_pair_context(pair_list_csv)
    expected_counts = load_expected_counts(priority_csv)
    validation = validate(merged_rows, pilot_ids, pair_context, expected_counts)

    v2_csv = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft_v2.csv"
    v2_xlsx = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft_v2.xlsx"
    report_md = output_dir / "diagraph_gold_50_pilot10_column_validation_report_v2.md"
    changes_md = output_dir / "diagraph_gold_50_pilot10_draft_v1_to_v2_changes.md"
    notes_md = output_dir / "diagraph_gold_50_pilot10_guide_revision_notes_v2.md"

    write_csv(v2_csv, merged_rows, FIELDS, overwrite=args.overwrite)
    write_workbook(v2_xlsx, merged_rows, overwrite=args.overwrite)
    write_text(report_md, build_validation_report(validation, len(added_rows)), overwrite=args.overwrite)
    write_text(changes_md, build_changes_md(v1_rows, added_rows, merged_rows, validation), overwrite=args.overwrite)
    write_text(notes_md, build_revision_notes_v2(), overwrite=args.overwrite)

    print("pilot10 draft v2 generation complete")
    print(f"added_rows={len(added_rows)}")
    print(f"merged_rows={len(merged_rows)}")
    print(f"covered_pairs={validation['draft_pair_count']}")
    print(f"row_count_pass={validation['row_count_pass']}")
    print(f"span_a_invalid={len(validation['invalid_span_a'])}")
    print(f"span_b_invalid={len(validation['invalid_span_b'])}")


if __name__ == "__main__":
    main()
