"""Replace F300V1-0127 pilot10 column annotations in v2 and generate v3."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from io_utils import artifact_path, ensure_can_write, read_csv, write_csv, write_text


FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
    "notes",
]

VALID_RELATION_TYPES = {
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
}
VALID_RELATION_STRENGTHS = {"strong", "medium", "weak"}
VALID_ALIGNMENT_DIRECTIONS = {"A_to_B", "B_to_A", "mutual"}
VALID_TERNARY = {"1", "0", "?"}

TARGET_ANNOTATION_ID = "F300V1-0127"
TARGET_PAIR_ID = "2715076"

CORRECTED_F300V1_0127_ROWS = [
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C01",
        "span_a": "大鹏与他是一母所生，故此有些亲处",
        "span_b": "这般比论",
        "relation_type": "coreference_or_demonstrative",
        "relation_strength": "medium",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B中的“这般比论”回指A中关于亲属关系的整段推理论证。",
    },
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C02",
        "span_a": "大鹏与他是一母所生",
        "span_b": "妖精的外甥",
        "relation_type": "analogy",
        "relation_strength": "strong",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B沿用A的亲属推理逻辑，将“大鹏与孔雀同母”讽刺性延展为“如来是妖精外甥”。",
    },
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C03",
        "span_a": "佛母",
        "span_b": "外甥",
        "relation_type": "analogy",
        "relation_strength": "strong",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "A中“佛母”建立宗教亲属关系，B以“外甥”改写并延展这一亲属关系。",
    },
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C04",
        "span_a": "大鹏",
        "span_b": "妖精",
        "relation_type": "semantic_substitution",
        "relation_strength": "medium",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B将A叙述中的“大鹏”贬称为“妖精”，形成评价性替换。",
    },
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C05",
        "span_a": "故此有些亲处",
        "span_b": "你还是妖精的外甥哩",
        "relation_type": "semantic_substitution",
        "relation_strength": "strong",
        "alignment_direction": "A_to_B",
        "is_core_column": "1",
        "supports_resonance": "1",
        "notes": "B把A的“有些亲处”具体化并讽刺化为“外甥”关系。",
    },
    {
        "annotation_id": TARGET_ANNOTATION_ID,
        "pair_id": TARGET_PAIR_ID,
        "column_id": "C06",
        "span_a": "佛母",
        "span_b": "你还是妖精的外甥哩",
        "relation_type": "analogy",
        "relation_strength": "medium",
        "alignment_direction": "A_to_B",
        "is_core_column": "0",
        "supports_resonance": "1",
        "notes": "A中的“佛母”亲属定位被B讽刺性改写为“外甥”关系；此栏辅助说明类比链条，不单独作为核心纵栏。",
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        default=str(artifact_path("formal_300_v1", "diagraph_gold_50")),
    )
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def normalize_column_id(column_id: str) -> str:
    value = str(column_id or "").strip()
    if re.fullmatch(r"C\d+", value):
        return f"C{int(value[1:]):02d}"
    return value


def row_sort_key(row: dict[str, str]) -> tuple[str, int, str]:
    match = re.fullmatch(r"C(\d+)", row["column_id"])
    number = int(match.group(1)) if match else 999
    return (row["annotation_id"], number, row["column_id"])


def load_pilot_ids(pilot10_path: Path) -> list[str]:
    pilot_ids: list[str] = []
    with pilot10_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        annotation_idx = header.index("annotation_id")
        for raw_row in reader:
            if not raw_row:
                continue
            annotation_id = raw_row[annotation_idx].strip()
            if annotation_id and annotation_id not in pilot_ids:
                pilot_ids.append(annotation_id)
    return pilot_ids[:10]


def load_pair_context(pair_list_path: Path) -> dict[str, dict[str, str]]:
    rows = read_csv(pair_list_path)
    return {row["annotation_id"]: row for row in rows}


def replace_target_rows(v2_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    old_rows = [row for row in v2_rows if row["annotation_id"] == TARGET_ANNOTATION_ID]
    kept_rows = [row for row in v2_rows if row["annotation_id"] != TARGET_ANNOTATION_ID]
    corrected_rows = [dict(row) for row in CORRECTED_F300V1_0127_ROWS]
    return sorted(kept_rows + corrected_rows, key=row_sort_key), old_rows


def write_workbook(path: Path, rows: list[dict[str, str]], *, overwrite: bool) -> None:
    output_path = ensure_can_write(path, overwrite=overwrite)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "pilot10_draft_v3"
    header_fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    widths = {
        "annotation_id": 16,
        "pair_id": 12,
        "column_id": 10,
        "span_a": 30,
        "span_b": 30,
        "relation_type": 28,
        "relation_strength": 16,
        "alignment_direction": 18,
        "is_core_column": 16,
        "supports_resonance": 20,
        "notes": 58,
    }
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
        sheet.column_dimensions[cell.column_letter].width = widths.get(field, 18)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = wrap
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def validate(
    rows: list[dict[str, str]],
    pilot_ids: list[str],
    pair_context: dict[str, dict[str, str]],
) -> dict[str, object]:
    draft_ids = {row["annotation_id"] for row in rows}
    row_counts: Counter[str] = Counter()
    core_counts: Counter[str] = Counter()
    invalid_span_a: list[str] = []
    invalid_span_b: list[str] = []
    invalid_relation_type: list[str] = []
    invalid_relation_strength: list[str] = []
    invalid_alignment: list[str] = []
    invalid_ternary: list[str] = []
    non_pilot_rows: list[str] = []

    for row in rows:
        annotation_id = row["annotation_id"]
        column_id = row["column_id"]
        row_counts[annotation_id] += 1
        if row["is_core_column"] == "1":
            core_counts[annotation_id] += 1
        if annotation_id not in pilot_ids:
            non_pilot_rows.append(f"{annotation_id}/{column_id}")

        context = pair_context.get(annotation_id)
        if not context:
            invalid_span_a.append(f"{annotation_id}/{column_id}: missing pair context")
            invalid_span_b.append(f"{annotation_id}/{column_id}: missing pair context")
            continue
        if not row["span_a"] or row["span_a"] not in context["turn_a"]:
            invalid_span_a.append(f"{annotation_id}/{column_id}: {row['span_a']}")
        if not row["span_b"] or row["span_b"] not in context["turn_b"]:
            invalid_span_b.append(f"{annotation_id}/{column_id}: {row['span_b']}")

        if row["relation_type"] not in VALID_RELATION_TYPES:
            invalid_relation_type.append(f"{annotation_id}/{column_id}: {row['relation_type']}")
        if row["relation_strength"] not in VALID_RELATION_STRENGTHS:
            invalid_relation_strength.append(f"{annotation_id}/{column_id}: {row['relation_strength']}")
        if row["alignment_direction"] not in VALID_ALIGNMENT_DIRECTIONS:
            invalid_alignment.append(f"{annotation_id}/{column_id}: {row['alignment_direction']}")
        if row["is_core_column"] not in VALID_TERNARY:
            invalid_ternary.append(f"{annotation_id}/{column_id}: is_core_column={row['is_core_column']}")
        if row["supports_resonance"] not in VALID_TERNARY:
            invalid_ternary.append(
                f"{annotation_id}/{column_id}: supports_resonance={row['supports_resonance']}"
            )

    missing_pairs = [annotation_id for annotation_id in pilot_ids if annotation_id not in draft_ids]
    no_row_pairs = [annotation_id for annotation_id in pilot_ids if row_counts[annotation_id] < 1]
    no_core_pairs = [annotation_id for annotation_id in pilot_ids if core_counts[annotation_id] < 1]
    over_five_pairs = [annotation_id for annotation_id, count in row_counts.items() if count > 5]

    return {
        "draft_pair_count": len(draft_ids),
        "draft_row_count": len(rows),
        "row_count_pass": len(rows) == 39,
        "all_pilot_pairs_covered": set(pilot_ids) == draft_ids,
        "missing_pairs": missing_pairs,
        "no_row_pairs": no_row_pairs,
        "no_core_pairs": no_core_pairs,
        "row_counts": row_counts,
        "core_counts": core_counts,
        "invalid_span_a": invalid_span_a,
        "invalid_span_b": invalid_span_b,
        "invalid_relation_type": invalid_relation_type,
        "invalid_relation_strength": invalid_relation_strength,
        "invalid_alignment": invalid_alignment,
        "invalid_ternary": invalid_ternary,
        "non_pilot_rows": non_pilot_rows,
        "over_five_pairs": over_five_pairs,
        "f300v1_0127_allow_six": row_counts[TARGET_ANNOTATION_ID] == 6,
    }


def yes_no(ok: bool) -> str:
    return "通过" if ok else "未通过"


def list_or_none(items: list[str], none_text: str = "无") -> list[str]:
    if not items:
        return [f"- {none_text}"]
    return [f"- {item}" for item in items]


def build_validation_report(validation: dict[str, object], old_rows: list[dict[str, str]]) -> str:
    row_counts = validation["row_counts"]
    core_counts = validation["core_counts"]
    assert isinstance(row_counts, Counter)
    assert isinstance(core_counts, Counter)

    invalid_span_a = validation["invalid_span_a"]
    invalid_span_b = validation["invalid_span_b"]
    invalid_relation_type = validation["invalid_relation_type"]
    invalid_relation_strength = validation["invalid_relation_strength"]
    invalid_alignment = validation["invalid_alignment"]
    invalid_ternary = validation["invalid_ternary"]
    non_pilot_rows = validation["non_pilot_rows"]
    missing_pairs = validation["missing_pairs"]
    no_row_pairs = validation["no_row_pairs"]
    no_core_pairs = validation["no_core_pairs"]
    over_five_pairs = validation["over_five_pairs"]
    assert isinstance(invalid_span_a, list)
    assert isinstance(invalid_span_b, list)
    assert isinstance(invalid_relation_type, list)
    assert isinstance(invalid_relation_strength, list)
    assert isinstance(invalid_alignment, list)
    assert isinstance(invalid_ternary, list)
    assert isinstance(non_pilot_rows, list)
    assert isinstance(missing_pairs, list)
    assert isinstance(no_row_pairs, list)
    assert isinstance(no_core_pairs, list)
    assert isinstance(over_five_pairs, list)

    lines = [
        "# diagraph_gold_50 pilot10 column validation report v3",
        "",
        "## 1. 基本情况",
        "",
        f"- v2 中被替换的 `{TARGET_ANNOTATION_ID}` 行数：{len(old_rows)}",
        f"- v3 中 `{TARGET_ANNOTATION_ID}` 修正后行数：{row_counts[TARGET_ANNOTATION_ID]}",
        f"- v3 总行数：{validation['draft_row_count']}",
        "- 目标总行数：39",
        f"- 总行数是否为 39：{yes_no(bool(validation['row_count_pass']))}",
        f"- v3 是否覆盖全部 10 个 pilot pair：{yes_no(bool(validation['all_pilot_pairs_covered']))}",
        "",
        "## 2. pair 覆盖情况",
        "",
    ]
    for annotation_id in sorted(row_counts):
        lines.append(f"- `{annotation_id}`: {row_counts[annotation_id]} 行，core={core_counts[annotation_id]}")
    lines.extend(
        [
            "",
            f"- 是否存在非 pilot10 annotation_id：{yes_no(not non_pilot_rows)}",
            *list_or_none(non_pilot_rows),
            f"- 是否缺失 pilot10 annotation_id：{yes_no(not missing_pairs)}",
            *list_or_none(missing_pairs),
            "",
            "## 3. core / row 约束",
            "",
            f"- 每个 pilot pair 至少有 1 行：{yes_no(not no_row_pairs)}",
            *list_or_none(no_row_pairs),
            f"- 每个 pilot pair 至少有 1 个 `is_core_column=1`：{yes_no(not no_core_pairs)}",
            *list_or_none(no_core_pairs),
            "",
            "## 4. span 校验",
            "",
            f"- span_a 校验：{yes_no(not invalid_span_a)}",
            *list_or_none(invalid_span_a, "全部通过"),
            f"- span_b 校验：{yes_no(not invalid_span_b)}",
            *list_or_none(invalid_span_b, "全部通过"),
            "",
            "## 5. 合法值域校验",
            "",
            f"- relation_type：{yes_no(not invalid_relation_type)}",
            *list_or_none(invalid_relation_type, "全部合法"),
            f"- relation_strength：{yes_no(not invalid_relation_strength)}",
            *list_or_none(invalid_relation_strength, "全部合法"),
            f"- alignment_direction：{yes_no(not invalid_alignment)}",
            *list_or_none(invalid_alignment, "全部合法"),
            f"- is_core_column / supports_resonance：{yes_no(not invalid_ternary)}",
            *list_or_none(invalid_ternary, "全部合法"),
            "",
            "## 6. 超过 5 行与 0127 特例",
            "",
        ]
    )
    if over_five_pairs:
        for annotation_id in over_five_pairs:
            lines.append(f"- `{annotation_id}`: {row_counts[annotation_id]} 行")
    else:
        lines.append("- 当前没有超过 5 行的样本")
    lines.append(
        f"- `F300V1-0127` 是否允许 6 行：{yes_no(bool(validation['f300v1_0127_allow_six']))}"
    )
    lines.extend(
        [
            "",
            "## 7. 人工确认项",
            "",
        ]
    )
    if invalid_span_a or invalid_span_b:
        lines.append("- 仍有 span 未通过严格字符串校验，需要人工确认或继续修正。")
    else:
        lines.append("- 严格字符串 span 校验全部通过；当前无机器层面的 span 人工确认项。")
        lines.append("- `F300V1-0127` 的 analogy / semantic_substitution 仍建议在二次复核中确认关系解释是否充分。")
    return "\n".join(lines) + "\n"


def build_changes_md(old_rows: list[dict[str, str]], new_rows: list[dict[str, str]], validation: dict[str, object]) -> str:
    invalid_spans = list(validation["invalid_span_a"]) + list(validation["invalid_span_b"])
    row_counts = validation["row_counts"]
    assert isinstance(row_counts, Counter)

    lines = [
        "# diagraph_gold_50 pilot10 draft v2 to v3 changes",
        "",
        "## 1. 修改范围",
        "",
        f"- 本次只替换 `{TARGET_ANNOTATION_ID}` 的 6 行 column annotation。",
        "- v2 文件未覆盖，原始空模板未修改。",
        "- v3 仍覆盖全部 10 个 pilot pair，总行数仍为 39。",
        "",
        "## 2. F300V1-0127 的具体变化",
        "",
        "- C01：将 v2 中反向放置的 `这般比论 -> 大鹏与他是一母所生，故此有些亲处` 修正为 `大鹏与他是一母所生，故此有些亲处 -> 这般比论`。",
        "- C02-C05：保留亲属推理链、类比、评价性替换和具体化替换，按修正版统一校验。",
        "- C06：移除同一话轮内部的 `如来 -> 你` 独立纵栏，改为跨 turn_a / turn_b 的 `佛母 -> 你还是妖精的外甥哩` 辅助类比纵栏。",
        "- `C6` 在输出中规范化为 `C06`，与 v2 既有列号格式保持一致。",
        "",
        "## 3. 替换前后行数",
        "",
        f"- v2 中 `{TARGET_ANNOTATION_ID}` 行数：{len(old_rows)}",
        f"- v3 中 `{TARGET_ANNOTATION_ID}` 行数：{len(new_rows)}",
        f"- v3 中 `{TARGET_ANNOTATION_ID}` 是否为 6 行：{yes_no(row_counts[TARGET_ANNOTATION_ID] == 6)}",
        "",
        "## 4. span 是否需要人工确认",
        "",
    ]
    if invalid_spans:
        lines.extend([f"- {item}" for item in invalid_spans])
    else:
        lines.append("- v3 严格字符串 span 校验全部通过；当前无机器层面的 span 人工确认项。")
    lines.extend(
        [
            "",
            "## 5. 容易混淆的 relation_type",
            "",
            "- `coreference_or_demonstrative` vs 同话轮内部指称：只有跨 turn_a / turn_b 才能作为独立纵栏。",
            "- `analogy` vs `semantic_substitution`：前者要求关系结构转移，后者强调可解释替换。",
            "- `semantic_substitution` vs 纯话题相关：必须指出哪两个成分形成替换。",
        ]
    )
    return "\n".join(lines) + "\n"


def build_revision_notes_v3() -> str:
    return """# diagraph_gold_50 pilot10 guide revision notes v3

## 1. 跨句纵栏必须跨越 turn_a / turn_b

跨句图谱纵栏必须连接 `turn_a` 与 `turn_b` 中的成分。同一话轮内部的指称关系不能单独作为跨句纵栏。

例如，如果“如来”和“你”都出现在 `turn_b` 中，则不能标为一个 A/B column。可以在 `notes` 中说明这是一种话轮内部指称或称呼转换，但不要作为独立纵栏。

## 2. A/B 说话权转换中的“你/我”

当 A 用“你”指向 B，而 B 在回应中用“我”承接同一参与者时，可以标为 `coreference_or_demonstrative`。标注时应保留原文中的“你/我”，不要改写成解释性实体名，并在 `notes` 中说明这是 speaker-role shift / deictic shift。

只有当两个 span 分别来自 `turn_a` 与 `turn_b`，且确实指向同一论元位置时，才可作为独立纵栏。

## 3. slot_filling 的适用范围

`slot_filling` 可以出现在定义问答、行动问答、名词性询问中。判断时看 A 是否提供待填槽位，B 是否给出可对位的填充值或完整补足。

如果 B 给出多个并列填充值，可以拆成多行；如果只是继续同一话题而没有明确槽位，则不要标为 `slot_filling`。

## 4. contrast 的适用范围

`contrast` 可用于时间、行动主体、评价立场的对照。关键是两端是否形成稳定可对位的反向关系。

单纯否定、单纯不同意或换话题，不自动构成 `contrast`。

## 5. repair 不等于普通否定

`repair` 必须有对前一行动、说法、判断或话语路线的修正、制止、纠偏或重构。普通否定如果没有对 A 的具体表达进行修复，就不应标为 `repair`。

## 6. analogy 必须有结构推理链

`analogy` 必须能看到关系结构的转移或延展。标注者应能说明：A 先建立了怎样的关系结构，B 又如何把这条结构转移、延展或反讽性映射到新的关系结论。

如果只能说“两个话轮都在谈同一主题”，但说不出结构链条，则不要标为 `analogy`。

## 7. semantic_substitution 与纯话题相关

`semantic_substitution` 必须有可解释替换，例如对同一主体、行动、命题或评价位置的改写。它不等于普通话题相关。

如果只能说明“它们谈的是同一件事”，但不能指出哪两个成分形成替换关系，则暂不标为 `semantic_substitution`。

## 8. hard 样本行数

hard 样本允许超过模板默认 5 行，但必须说明新增行为什么支持跨句共鸣。`F300V1-0127` 可保留 6 行，其中非 core 的辅助纵栏应在 `notes` 中说明其辅助性质。
"""


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    expected_dir = artifact_path("formal_300_v1", "diagraph_gold_50").resolve()
    if output_dir.resolve() != expected_dir:
        raise SystemExit(f"Outputs must stay under {expected_dir}")

    v2_csv = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft_v2.csv"
    pair_list_csv = output_dir / "diagraph_gold_50_pair_list.csv"
    pilot10_csv = output_dir / "diagraph_gold_50_pilot10_list.csv"
    required = [v2_csv, pair_list_csv, pilot10_csv]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit(f"Required input files are missing: {missing}")

    v2_rows = read_csv(v2_csv)
    v3_rows, old_target_rows = replace_target_rows(v2_rows)
    new_target_rows = [row for row in v3_rows if row["annotation_id"] == TARGET_ANNOTATION_ID]

    pilot_ids = load_pilot_ids(pilot10_csv)
    pair_context = load_pair_context(pair_list_csv)
    validation = validate(v3_rows, pilot_ids, pair_context)

    v3_csv = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft_v3.csv"
    v3_xlsx = output_dir / "diagraph_gold_50_pilot10_column_annotation_draft_v3.xlsx"
    report_md = output_dir / "diagraph_gold_50_pilot10_column_validation_report_v3.md"
    changes_md = output_dir / "diagraph_gold_50_pilot10_draft_v2_to_v3_changes.md"
    notes_md = output_dir / "diagraph_gold_50_pilot10_guide_revision_notes_v3.md"

    write_csv(v3_csv, v3_rows, FIELDS, overwrite=args.overwrite)
    write_workbook(v3_xlsx, v3_rows, overwrite=args.overwrite)
    write_text(report_md, build_validation_report(validation, old_target_rows), overwrite=args.overwrite)
    write_text(changes_md, build_changes_md(old_target_rows, new_target_rows, validation), overwrite=args.overwrite)
    write_text(notes_md, build_revision_notes_v3(), overwrite=args.overwrite)

    print("pilot10 draft v3 generation complete")
    print(f"replaced_annotation_id={TARGET_ANNOTATION_ID}")
    print(f"old_target_rows={len(old_target_rows)}")
    print(f"new_target_rows={len(new_target_rows)}")
    print(f"merged_rows={len(v3_rows)}")
    print(f"covered_pairs={validation['draft_pair_count']}")
    print(f"row_count_pass={validation['row_count_pass']}")
    print(f"span_a_invalid={len(validation['invalid_span_a'])}")
    print(f"span_b_invalid={len(validation['invalid_span_b'])}")
    print(f"value_domain_invalid={sum(len(validation[key]) for key in ['invalid_relation_type', 'invalid_relation_strength', 'invalid_alignment', 'invalid_ternary'])}")


if __name__ == "__main__":
    main()
