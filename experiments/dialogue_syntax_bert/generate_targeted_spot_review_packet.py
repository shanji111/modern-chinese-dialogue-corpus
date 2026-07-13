from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
BASE_DIR = ROOT / "artifacts" / "formal_300_v1" / "diagraph_gold_50"
FULL_DIR = BASE_DIR / "full_gold_candidate"
SANITY_DIR = FULL_DIR / "final_sanity_check"
OUTPUT_DIR = SANITY_DIR / "targeted_spot_review"

ALL_ROWS_PATH = FULL_DIR / "full_diagraph_gold_50_column_reviewed_all_rows.csv"
ACTIVE_PATH = FULL_DIR / "full_diagraph_gold_50_column_gold_candidate_active.csv"
SPOT_REVIEW_LIST_PATH = SANITY_DIR / "full_diagraph_gold_50_final_spot_review_list.csv"
MERGE_REPORT_PATH = FULL_DIR / "full_diagraph_gold_50_merge_validation_report.md"
DISTRIBUTION_SUMMARY_PATH = FULL_DIR / "full_diagraph_gold_50_distribution_summary.md"
PAIR_LIST_PATH = BASE_DIR / "diagraph_gold_50_pair_list.csv"
GUIDE_PATH = BASE_DIR / "diagraph_gold_50_annotation_guide_v2.md"

PACKET_MD_PATH = OUTPUT_DIR / "targeted_spot_review_packet.md"
PACKET_XLSX_PATH = OUTPUT_DIR / "targeted_spot_review_packet.xlsx"
DECISION_TEMPLATE_CSV_PATH = OUTPUT_DIR / "targeted_spot_review_decisions_template.csv"
DECISION_TEMPLATE_XLSX_PATH = OUTPUT_DIR / "targeted_spot_review_decisions_template.xlsx"
GUIDE_MD_PATH = OUTPUT_DIR / "targeted_spot_review_guide.md"
VALIDATION_MD_PATH = OUTPUT_DIR / "targeted_spot_review_validation_report.md"

TARGET_IDS = [
    "F300V1-0127",
    "F300V1-0220",
    "F300V1-0287",
    "F300V1-0214",
    "F300V1-0111",
    "F300V1-0254",
]
TARGET_SET = set(TARGET_IDS)

VALID_RELATION_TYPES = {
    "lexical_reproduction",
    "syntactic_parallelism",
    "semantic_substitution",
    "coreference_or_demonstrative",
    "slot_filling",
    "short_answer",
    "contrast",
    "repair",
    "analogy",
    "pragmatic_function",
    "punctuation_or_modal",
    "other",
}
VALID_STRENGTHS = {"strong", "medium", "weak"}
VALID_DIRECTIONS = {"A_to_B", "B_to_A", "mutual"}
VALID_BINARY = {"0", "1"}
DECISION_OPTIONS = ["keep", "revise", "delete", "unsure"]

PAIR_FIELDS = [
    "annotation_id",
    "pair_id",
    "source",
    "dataset_name",
    "sample_stratum",
    "turn_a",
    "turn_b",
    "resonance_present",
    "label_reproduction",
    "label_parallelism",
    "label_selective_reuse",
    "label_repair",
    "label_contrast",
    "label_analogy_candidate",
    "evidence_span_a",
    "evidence_span_b",
    "annotator_note",
    "rule_any_positive",
    "bert_prob",
    "hybrid_pred",
]

ACTIVE_FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "span_a",
    "span_b",
    "relation_type",
    "relation_strength",
    "alignment_direction",
    "is_core_column",
    "supports_resonance",
    "notes",
    "reviewer_decision",
    "reviewer_note",
    "batch",
    "source",
    "dataset_name",
    "difficulty_level",
    "sample_stratum",
]

DECISION_TEMPLATE_FIELDS = [
    "annotation_id",
    "pair_id",
    "column_id",
    "current_relation_type",
    "current_relation_strength",
    "current_is_core_column",
    "current_supports_resonance",
    "suggested_decision",
    "suggested_relation_type",
    "suggested_relation_strength",
    "suggested_is_core_column",
    "suggested_supports_resonance",
    "reviewer_note",
    "freeze_impact",
]

SPOT_REVIEW_GUIDE = {
    "F300V1-0127": [
        "只检查 analogy 主链是否真的成立，不重新扩展新的类比栏。",
        "重点核对 C06 作为 non-core auxiliary analogy chain 是否仍然合理，避免它与主链重复承担同一功能。",
        "若要调整，只围绕现有 C02/C03/C06 的分工与边界，不改 candidate 结构。",
    ],
    "F300V1-0220": [
        "只检查 analogy 主链是否成立，确认 C01/C02 不是把同一结构链机械拆成两栏。",
        "不要重新扩展已删除的行动者位辅助栏；本轮只复核 active candidate。",
        "若结构推理链说不清，应在 decision template 中标 revise/unsure，而不是追加新栏。",
    ],
    "F300V1-0287": [
        "重点检查 semantic_substitution 是否仍有过度概括风险，尤其是 C03/C05 是否真有明确替换位。",
        "pragmatic_function 仅在确属回应性立场表达时保留，不要把政策话题相关误收成纵栏。",
        "如果辅助栏多于必要范围，应优先考虑降格或删除，而不是继续增殖。",
    ],
    "F300V1-0214": [
        "检查单栏 pragmatic_function 是否足以支撑该 pair，避免把纯论坛语境延续误当成稳定映射。",
        "若唯一核心栏过弱，应明确记录风险，而不是补造缺乏证据的新栏。",
        "特别注意 short_answer / pragmatic_function / topic-relatedness 的边界。",
    ],
    "F300V1-0111": [
        "重点检查“某些球迷→破车迷”是否可作为论坛语境中的标签替换，而不是纯情绪性压缩。",
        "若认定它只是同话题标签化评价，应考虑 revise/unsure，而不是默认 semantic_substitution 成立。",
        "不重新打开已删除栏，只复核现存 active column 的合法性。",
    ],
    "F300V1-0254": [
        "重点检查 core column 是否偏满，尤其看 C02/C05 是否至少有一栏更适合作为 auxiliary。",
        "repair 与 contrast 分工要清楚：修正前一行动与重新分配行动主体不是同一回事。",
        "如果主链仍成立但并列 core 过多，可考虑降格而不是整栏删除。",
    ],
}


THIN_GRAY = Side(style="thin", color="D9D9D9")
THIN_BLUE = Side(style="thin", color="9CC2E5")
THIN_LIGHT_BLUE = Side(style="thin", color="B8CCE4")
BORDER_GRAY = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
BORDER_BLUE = Border(left=THIN_BLUE, right=THIN_BLUE, top=THIN_BLUE, bottom=THIN_BLUE)
BORDER_LIGHT_BLUE = Border(
    left=THIN_LIGHT_BLUE,
    right=THIN_LIGHT_BLUE,
    top=THIN_LIGHT_BLUE,
    bottom=THIN_LIGHT_BLUE,
)


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: Iterable[Dict[str, str]], fieldnames: List[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def ensure_inputs_exist() -> None:
    required = [
        ALL_ROWS_PATH,
        ACTIVE_PATH,
        SPOT_REVIEW_LIST_PATH,
        MERGE_REPORT_PATH,
        DISTRIBUTION_SUMMARY_PATH,
        PAIR_LIST_PATH,
        GUIDE_PATH,
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required inputs: {missing}")


def sort_columns(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    def column_order(row: Dict[str, str]) -> tuple[int, str]:
        column_id = row["column_id"]
        digits = "".join(ch for ch in column_id if ch.isdigit())
        return (int(digits) if digits else 0, column_id)

    return sorted(rows, key=column_order)


def pair_label_summary(pair: Dict[str, str]) -> str:
    return " | ".join(
        [
            f"reproduction={pair['label_reproduction']}",
            f"parallelism={pair['label_parallelism']}",
            f"selective_reuse={pair['label_selective_reuse']}",
            f"repair={pair['label_repair']}",
            f"contrast={pair['label_contrast']}",
            f"analogy_candidate={pair['label_analogy_candidate']}",
        ]
    )


def row_key(row: Dict[str, str]) -> str:
    return f"{row['annotation_id']}/{row['column_id']}"


def escape_md(text: str) -> str:
    return str(text or "").replace("|", "\\|").replace("\n", "<br>")


def make_freeze_impact(row: Dict[str, str], spot_row: Dict[str, str]) -> str:
    priority = spot_row["priority"]
    if row["is_core_column"] == "1" and priority == "high":
        return "high: revising this core column may affect freeze readiness"
    if row["is_core_column"] == "1":
        return "medium: revising this core column may change the retained main chain"
    if priority == "high":
        return "medium: auxiliary revision may still alter the risk profile of this pair"
    return "low: mainly affects auxiliary interpretation"


def build_focus_text(annotation_id: str, spot_row: Dict[str, str]) -> str:
    lines = [
        f"risk_type={spot_row['risk_type']}",
        f"reason={spot_row['reason']}",
        f"suggested_action={spot_row['suggested_action']}",
    ]
    lines.extend(SPOT_REVIEW_GUIDE[annotation_id])
    return "\n".join(f"{idx}. {line}" for idx, line in enumerate(lines, start=1))


def build_review_items(
    pair_rows: List[Dict[str, str]],
    active_rows: List[Dict[str, str]],
    spot_review_rows: List[Dict[str, str]],
) -> List[Dict[str, object]]:
    pair_map = {row["annotation_id"]: row for row in pair_rows}
    active_subset = [row for row in active_rows if row["annotation_id"] in TARGET_SET]
    active_by_annotation: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in active_subset:
        active_by_annotation[row["annotation_id"]].append(row)

    spot_map = {row["annotation_id"]: row for row in spot_review_rows}
    items: List[Dict[str, object]] = []
    for annotation_id in TARGET_IDS:
        pair = pair_map.get(annotation_id)
        spot_row = spot_map.get(annotation_id)
        columns = sort_columns(active_by_annotation.get(annotation_id, []))
        if pair is None:
            raise ValueError(f"Pair metadata missing for {annotation_id}")
        if spot_row is None:
            raise ValueError(f"Spot review row missing for {annotation_id}")
        if not columns:
            raise ValueError(f"Active candidate columns missing for {annotation_id}")
        items.append(
            {
                "annotation_id": annotation_id,
                "pair": pair,
                "spot": spot_row,
                "columns": columns,
                "focus_lines": SPOT_REVIEW_GUIDE[annotation_id],
                "focus_text": build_focus_text(annotation_id, spot_row),
            }
        )
    return items


def build_decision_template_rows(items: List[Dict[str, object]]) -> List[Dict[str, str]]:
    template_rows: List[Dict[str, str]] = []
    for item in items:
        pair = item["pair"]  # type: ignore[assignment]
        spot = item["spot"]  # type: ignore[assignment]
        for row in item["columns"]:  # type: ignore[index]
            template_rows.append(
                {
                    "annotation_id": row["annotation_id"],
                    "pair_id": pair["pair_id"],
                    "column_id": row["column_id"],
                    "current_relation_type": row["relation_type"],
                    "current_relation_strength": row["relation_strength"],
                    "current_is_core_column": row["is_core_column"],
                    "current_supports_resonance": row["supports_resonance"],
                    "suggested_decision": "",
                    "suggested_relation_type": "",
                    "suggested_relation_strength": "",
                    "suggested_is_core_column": "",
                    "suggested_supports_resonance": "",
                    "reviewer_note": "",
                    "freeze_impact": make_freeze_impact(row, spot),
                }
            )
    return template_rows


def build_packet_markdown(items: List[Dict[str, object]]) -> str:
    lines = [
        "# targeted spot review packet",
        "",
        "本包只抽取 final sanity check 指定的 6 个 active candidate 样本，不新增 column，不修改 candidate 本体。",
        "",
    ]
    for item in items:
        pair = item["pair"]  # type: ignore[assignment]
        spot = item["spot"]  # type: ignore[assignment]
        columns = item["columns"]  # type: ignore[assignment]
        focus_lines = item["focus_lines"]  # type: ignore[assignment]
        lines.append(f"## {item['annotation_id']}")
        lines.append("")
        lines.append(f"- pair_id: `{pair['pair_id']}`")
        lines.append(f"- batch: `{columns[0]['batch']}`")
        lines.append(f"- source / dataset: {pair['source']} / {pair['dataset_name']}")
        lines.append(f"- difficulty_level: `{columns[0]['difficulty_level']}`")
        lines.append(f"- pair-level labels: {pair_label_summary(pair)}")
        lines.append(f"- evidence_span_a: {pair['evidence_span_a']}")
        lines.append(f"- evidence_span_b: {pair['evidence_span_b']}")
        lines.append(f"- spot review reason: {spot['reason']}")
        lines.append(f"- spot review priority: `{spot['priority']}`")
        lines.append("")
        lines.append("**Turn A**")
        lines.append(pair["turn_a"])
        lines.append("")
        lines.append("**Turn B**")
        lines.append(pair["turn_b"])
        lines.append("")
        lines.append("**Spot Review Focus**")
        for focus in focus_lines:
            lines.append(f"- {focus}")
        lines.append("")
        lines.append(
            "| column_id | span_a | span_b | relation_type | relation_strength | alignment_direction | is_core_column | supports_resonance | reviewer_decision | reviewer_note | notes |"
        )
        lines.append(
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |"
        )
        for column in columns:
            lines.append(
                "| "
                + " | ".join(
                    [
                        escape_md(column["column_id"]),
                        escape_md(column["span_a"]),
                        escape_md(column["span_b"]),
                        escape_md(column["relation_type"]),
                        escape_md(column["relation_strength"]),
                        escape_md(column["alignment_direction"]),
                        escape_md(column["is_core_column"]),
                        escape_md(column["supports_resonance"]),
                        escape_md(column["reviewer_decision"]),
                        escape_md(column["reviewer_note"]),
                        escape_md(column["notes"]),
                    ]
                )
                + " |"
            )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def build_guide_markdown(items: List[Dict[str, object]]) -> str:
    lines = [
        "# targeted spot review guide",
        "",
        "本轮只对 6 个高风险样本做 targeted spot review。范围限定为 active candidate 的现存 columns，不新增 column，不回滚 delete 历史，不改写 candidate 文件。",
        "",
        "## General Rules",
        "",
        "- 只复核 active columns，不自动扩展已删除栏。",
        "- `span_a` 必须来自 `turn_a`，`span_b` 必须来自 `turn_b`。",
        "- `analogy` 必须能说清 A 的关系结构如何转移、延展或反讽映射到 B。",
        "- `semantic_substitution` 必须有明确替换位，不能只是普通话题相关。",
        "- `pragmatic_function` 只用于语用性回应、确认请求或解释性回应。",
        "- 如果有疑问，优先在 decision template 里填 `revise` 或 `unsure`，而不是新增新栏。",
        "",
    ]
    for item in items:
        spot = item["spot"]  # type: ignore[assignment]
        lines.append(f"## {item['annotation_id']}")
        lines.append("")
        lines.append(f"- priority: `{spot['priority']}`")
        lines.append(f"- risk_type: `{spot['risk_type']}`")
        lines.append(f"- reason: {spot['reason']}")
        lines.append(f"- suggested_action: {spot['suggested_action']}")
        for focus in item["focus_lines"]:  # type: ignore[index]
            lines.append(f"- {focus}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def validate(
    items: List[Dict[str, object]],
    all_rows: List[Dict[str, str]],
    pair_rows: List[Dict[str, str]],
    template_rows: List[Dict[str, str]],
    input_mtimes_before: Dict[str, int],
) -> Dict[str, object]:
    pair_map = {row["annotation_id"]: row for row in pair_rows}
    active_ids = [item["annotation_id"] for item in items]
    active_keys = {row_key(row) for item in items for row in item["columns"]}  # type: ignore[index]
    template_keys = {f"{row['annotation_id']}/{row['column_id']}" for row in template_rows}
    per_pair_row_counts: Dict[str, int] = {}
    per_pair_core_counts: Dict[str, int] = {}
    span_a_failures: List[str] = []
    span_b_failures: List[str] = []
    invalid_relation_rows: List[str] = []
    invalid_strength_rows: List[str] = []
    invalid_direction_rows: List[str] = []
    invalid_binary_rows: List[str] = []
    non_active_or_new_columns: List[str] = sorted(template_keys - active_keys)
    missing_template_columns: List[str] = sorted(active_keys - template_keys)

    for item in items:
        pair = item["pair"]  # type: ignore[assignment]
        columns = item["columns"]  # type: ignore[assignment]
        per_pair_row_counts[item["annotation_id"]] = len(columns)
        per_pair_core_counts[item["annotation_id"]] = sum(
            1 for row in columns if row["is_core_column"] == "1"
        )
        for row in columns:
            key = row_key(row)
            if row["span_a"] not in pair["turn_a"]:
                span_a_failures.append(key)
            if row["span_b"] not in pair["turn_b"]:
                span_b_failures.append(key)
            if row["relation_type"] not in VALID_RELATION_TYPES:
                invalid_relation_rows.append(key)
            if row["relation_strength"] not in VALID_STRENGTHS:
                invalid_strength_rows.append(key)
            if row["alignment_direction"] not in VALID_DIRECTIONS:
                invalid_direction_rows.append(key)
            if row["is_core_column"] not in VALID_BINARY or row["supports_resonance"] not in VALID_BINARY:
                invalid_binary_rows.append(key)

    all_row_counts = Counter()
    for row in all_rows:
        if row["annotation_id"] in TARGET_SET:
            all_row_counts[row["annotation_id"]] += 1

    input_mtimes_after = {
        "all_rows": ALL_ROWS_PATH.stat().st_mtime_ns,
        "active": ACTIVE_PATH.stat().st_mtime_ns,
        "spot_review_list": SPOT_REVIEW_LIST_PATH.stat().st_mtime_ns,
        "pair_list": PAIR_LIST_PATH.stat().st_mtime_ns,
        "guide_v2": GUIDE_PATH.stat().st_mtime_ns,
    }
    input_unchanged = input_mtimes_before == input_mtimes_after

    return {
        "target_ids_match": active_ids == TARGET_IDS,
        "unique_ids": active_ids,
        "all_from_active_subset": template_keys.issubset(active_keys),
        "all_in_pair_list": all(annotation_id in pair_map for annotation_id in active_ids),
        "template_matches_active": not non_active_or_new_columns and not missing_template_columns,
        "non_active_or_new_columns": non_active_or_new_columns,
        "missing_template_columns": missing_template_columns,
        "per_pair_row_counts": per_pair_row_counts,
        "per_pair_core_counts": per_pair_core_counts,
        "span_a_failures": span_a_failures,
        "span_b_failures": span_b_failures,
        "invalid_relation_rows": invalid_relation_rows,
        "invalid_strength_rows": invalid_strength_rows,
        "invalid_direction_rows": invalid_direction_rows,
        "invalid_binary_rows": invalid_binary_rows,
        "all_row_counts": all_row_counts,
        "active_row_count": len(active_keys),
        "input_unchanged": input_unchanged,
        "input_mtimes_before": input_mtimes_before,
        "input_mtimes_after": input_mtimes_after,
        "all_rows_total_for_targets": sum(all_row_counts.values()),
    }


def build_validation_markdown(validation: Dict[str, object]) -> str:
    lines = [
        "# targeted spot review validation report",
        "",
        "## Scope",
        "",
        f"- target annotation_id count: {len(TARGET_IDS)}",
        f"- targeted annotation_ids: {', '.join(validation['unique_ids'])}",
        f"- active candidate rows in packet/template: {validation['active_row_count']}",
        f"- all reviewed rows across these 6 ids (including deleted history): {validation['all_rows_total_for_targets']}",
        "",
        "## Checks",
        "",
        f"- only the 6 specified annotation_ids: {'PASS' if validation['target_ids_match'] else 'FAIL'}",
        f"- all rows come from active candidate subset: {'PASS' if validation['all_from_active_subset'] else 'FAIL'}",
        f"- no new columns introduced and no active columns omitted: {'PASS' if validation['template_matches_active'] else 'FAIL'}",
        f"- candidate input files unchanged during generation: {'PASS' if validation['input_unchanged'] else 'FAIL'}",
        f"- span_a in turn_a: {'PASS' if not validation['span_a_failures'] else 'FAIL'}",
        f"- span_b in turn_b: {'PASS' if not validation['span_b_failures'] else 'FAIL'}",
        f"- relation_type legal: {'PASS' if not validation['invalid_relation_rows'] else 'FAIL'}",
        f"- relation_strength legal: {'PASS' if not validation['invalid_strength_rows'] else 'FAIL'}",
        f"- alignment_direction legal: {'PASS' if not validation['invalid_direction_rows'] else 'FAIL'}",
        f"- is_core_column / supports_resonance legal: {'PASS' if not validation['invalid_binary_rows'] else 'FAIL'}",
        "",
        "## Per-Pair Counts",
        "",
        "| annotation_id | active_rows | active_core_columns | all_rows_in_review_history |",
        "| --- | ---: | ---: | ---: |",
    ]
    for annotation_id in TARGET_IDS:
        lines.append(
            f"| {annotation_id} | {validation['per_pair_row_counts'][annotation_id]} | "
            f"{validation['per_pair_core_counts'][annotation_id]} | {validation['all_row_counts'][annotation_id]} |"
        )

    if validation["non_active_or_new_columns"]:
        lines.extend(
            [
                "",
                "## Unexpected Extra Columns",
                "",
                *[f"- {key}" for key in validation["non_active_or_new_columns"]],
            ]
        )
    if validation["missing_template_columns"]:
        lines.extend(
            [
                "",
                "## Missing Active Columns",
                "",
                *[f"- {key}" for key in validation["missing_template_columns"]],
            ]
        )
    if validation["span_a_failures"]:
        lines.extend(
            [
                "",
                "## span_a Failures",
                "",
                *[f"- {key}" for key in validation["span_a_failures"]],
            ]
        )
    if validation["span_b_failures"]:
        lines.extend(
            [
                "",
                "## span_b Failures",
                "",
                *[f"- {key}" for key in validation["span_b_failures"]],
            ]
        )

    lines.extend(
        [
            "",
            "## Constraint Audit",
            "",
            "- candidate file modification: none",
            "- BERT training or inference: none",
            "- gold_v1 / gold_v1_binary modification: none",
            "- train/dev/test split modification: none",
            "- formal corpus.db read/write: none",
            "- website routing / deployment: none",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def set_label_cell(ws, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor="DDEBF7")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER_LIGHT_BLUE


def set_value_cell(ws, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER_GRAY


def merge_label_value(ws, label_cell: str, value_range: str, label: str, value: str) -> None:
    set_label_cell(ws, label_cell, label)
    ws.merge_cells(value_range)
    start_cell = value_range.split(":")[0]
    set_value_cell(ws, start_cell, value)


def autosize(ws, widths: Dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_table_header(ws, row_idx: int, last_col_idx: int) -> None:
    for col_idx in range(1, last_col_idx + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_BLUE


def style_table_body(ws, start_row: int, end_row: int, last_col_idx: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, last_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER_GRAY


def write_packet_workbook(items: List[Dict[str, object]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    index_ws = workbook.create_sheet("Index")
    index_ws.sheet_view.showGridLines = False
    index_headers = [
        "annotation_id",
        "pair_id",
        "batch",
        "difficulty_level",
        "priority",
        "active_column_count",
        "core_column_count",
        "risk_type",
        "spot_review_focus",
    ]
    index_ws.append(index_headers)
    style_table_header(index_ws, 1, len(index_headers))
    for item in items:
        pair = item["pair"]  # type: ignore[assignment]
        spot = item["spot"]  # type: ignore[assignment]
        columns = item["columns"]  # type: ignore[assignment]
        index_ws.append(
            [
                item["annotation_id"],
                pair["pair_id"],
                columns[0]["batch"],
                columns[0]["difficulty_level"],
                spot["priority"],
                len(columns),
                sum(1 for row in columns if row["is_core_column"] == "1"),
                spot["risk_type"],
                " / ".join(item["focus_lines"]),  # type: ignore[index]
            ]
        )
    style_table_body(index_ws, 2, index_ws.max_row, len(index_headers))
    index_ws.freeze_panes = "A2"
    autosize(
        index_ws,
        {
            "A": 16,
            "B": 12,
            "C": 14,
            "D": 12,
            "E": 10,
            "F": 16,
            "G": 16,
            "H": 28,
            "I": 60,
        },
    )

    for item in items:
        pair = item["pair"]  # type: ignore[assignment]
        spot = item["spot"]  # type: ignore[assignment]
        columns = item["columns"]  # type: ignore[assignment]
        ws = workbook.create_sheet(item["annotation_id"])
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:K1")
        ws["A1"] = f"Targeted Spot Review Packet | {item['annotation_id']}"
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].border = BORDER_BLUE

        set_label_cell(ws, "A2", "annotation_id")
        set_value_cell(ws, "B2", item["annotation_id"])
        set_label_cell(ws, "D2", "pair_id")
        set_value_cell(ws, "E2", pair["pair_id"])
        set_label_cell(ws, "G2", "batch")
        set_value_cell(ws, "H2", columns[0]["batch"])
        set_label_cell(ws, "J2", "difficulty_level")
        set_value_cell(ws, "K2", columns[0]["difficulty_level"])

        set_label_cell(ws, "A3", "source")
        set_value_cell(ws, "B3", pair["source"])
        set_label_cell(ws, "D3", "dataset_name")
        set_value_cell(ws, "E3", pair["dataset_name"])
        set_label_cell(ws, "G3", "priority")
        set_value_cell(ws, "H3", spot["priority"])
        set_label_cell(ws, "J3", "risk_type")
        set_value_cell(ws, "K3", spot["risk_type"])

        merge_label_value(ws, "A4", "B4:K4", "pair_level_labels", pair_label_summary(pair))
        merge_label_value(ws, "A5", "B5:K5", "evidence_span_a", pair["evidence_span_a"])
        merge_label_value(ws, "A6", "B6:K6", "evidence_span_b", pair["evidence_span_b"])
        merge_label_value(ws, "A7", "B7:K8", "spot_review_focus", item["focus_text"])  # type: ignore[arg-type]
        merge_label_value(ws, "A9", "B9:K12", "turn_a", pair["turn_a"])
        merge_label_value(ws, "A13", "B13:K16", "turn_b", pair["turn_b"])

        header_row = 18
        headers = [
            "column_id",
            "span_a",
            "span_b",
            "relation_type",
            "relation_strength",
            "alignment_direction",
            "is_core_column",
            "supports_resonance",
            "notes",
            "reviewer_decision",
            "reviewer_note",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=header)
        style_table_header(ws, header_row, len(headers))

        for row_offset, row in enumerate(columns, start=1):
            values = [
                row["column_id"],
                row["span_a"],
                row["span_b"],
                row["relation_type"],
                row["relation_strength"],
                row["alignment_direction"],
                row["is_core_column"],
                row["supports_resonance"],
                row["notes"],
                row["reviewer_decision"],
                row["reviewer_note"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=header_row + row_offset, column=col_idx, value=value)
        if columns:
            style_table_body(ws, header_row + 1, header_row + len(columns), len(headers))
        ws.freeze_panes = f"A{header_row + 1}"
        autosize(
            ws,
            {
                "A": 10,
                "B": 30,
                "C": 30,
                "D": 24,
                "E": 16,
                "F": 18,
                "G": 12,
                "H": 16,
                "I": 42,
                "J": 16,
                "K": 34,
            },
        )

    workbook.save(PACKET_XLSX_PATH)
    load_workbook(PACKET_XLSX_PATH)


def write_decision_template_workbook(rows: List[Dict[str, str]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "DecisionsTemplate"
    ws.sheet_view.showGridLines = False

    ws.append(DECISION_TEMPLATE_FIELDS)
    style_table_header(ws, 1, len(DECISION_TEMPLATE_FIELDS))
    for row in rows:
        ws.append([row[field] for field in DECISION_TEMPLATE_FIELDS])
    if rows:
        style_table_body(ws, 2, ws.max_row, len(DECISION_TEMPLATE_FIELDS))
    ws.freeze_panes = "A2"
    autosize(
        ws,
        {
            "A": 16,
            "B": 12,
            "C": 10,
            "D": 24,
            "E": 18,
            "F": 18,
            "G": 22,
            "H": 18,
            "I": 24,
            "J": 24,
            "K": 22,
            "L": 28,
            "M": 36,
            "N": 48,
        },
    )

    decision_col_letter = "H"
    decision_validation = DataValidation(
        type="list",
        formula1='"keep,revise,delete,unsure"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_validation)
    if ws.max_row >= 2:
        decision_validation.add(f"{decision_col_letter}2:{decision_col_letter}{ws.max_row}")

    workbook.save(DECISION_TEMPLATE_XLSX_PATH)
    load_workbook(DECISION_TEMPLATE_XLSX_PATH)


def main() -> None:
    ensure_inputs_exist()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    input_mtimes_before = {
        "all_rows": ALL_ROWS_PATH.stat().st_mtime_ns,
        "active": ACTIVE_PATH.stat().st_mtime_ns,
        "spot_review_list": SPOT_REVIEW_LIST_PATH.stat().st_mtime_ns,
        "pair_list": PAIR_LIST_PATH.stat().st_mtime_ns,
        "guide_v2": GUIDE_PATH.stat().st_mtime_ns,
    }

    pair_rows = read_csv_dicts(PAIR_LIST_PATH)
    active_rows = read_csv_dicts(ACTIVE_PATH)
    all_rows = read_csv_dicts(ALL_ROWS_PATH)
    spot_review_rows = read_csv_dicts(SPOT_REVIEW_LIST_PATH)

    if [row["annotation_id"] for row in spot_review_rows] != TARGET_IDS:
        raise ValueError("Spot review list does not match the required 6 annotation_ids in order.")

    review_items = build_review_items(pair_rows, active_rows, spot_review_rows)
    template_rows = build_decision_template_rows(review_items)

    write_csv(DECISION_TEMPLATE_CSV_PATH, template_rows, DECISION_TEMPLATE_FIELDS)
    PACKET_MD_PATH.write_text(build_packet_markdown(review_items), encoding="utf-8")
    GUIDE_MD_PATH.write_text(build_guide_markdown(review_items), encoding="utf-8")
    write_packet_workbook(review_items)
    write_decision_template_workbook(template_rows)

    validation = validate(
        items=review_items,
        all_rows=all_rows,
        pair_rows=pair_rows,
        template_rows=template_rows,
        input_mtimes_before=input_mtimes_before,
    )
    VALIDATION_MD_PATH.write_text(build_validation_markdown(validation), encoding="utf-8")

    print(f"target_pairs={len(review_items)}")
    print(f"active_rows={validation['active_row_count']}")
    print(f"packet_md={PACKET_MD_PATH}")
    print(f"packet_xlsx={PACKET_XLSX_PATH}")
    print(f"decision_csv={DECISION_TEMPLATE_CSV_PATH}")
    print(f"decision_xlsx={DECISION_TEMPLATE_XLSX_PATH}")
    print(f"guide_md={GUIDE_MD_PATH}")
    print(f"validation_md={VALIDATION_MD_PATH}")
    print(f"inputs_unchanged={validation['input_unchanged']}")


if __name__ == "__main__":
    main()
